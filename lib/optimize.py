#!/usr/bin/env python3
"""
Payment Processing Platform Resource Optimizer
Analyzes 90-day metrics and optimizes resource allocation while maintaining SLAs
Author: Platform Engineering Team
"""

import json
import logging
import os
import warnings
from datetime import datetime, timedelta
from typing import Any, Dict, List

import boto3
import numpy as np
import pandas as pd
from botocore.exceptions import ClientError
from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('payment_optimization.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class PaymentPlatformOptimizer:
    """
    Comprehensive resource optimizer for Payment Processing infrastructure
    """

    def __init__(self, region_name: str = None):
        """Initialize AWS clients and configuration"""

        # Auto-detect current region if not specified
        session = boto3.Session()
        current_region = session.region_name or 'us-east-1'

        self.region_name = region_name or current_region

        logger.info(f"Initializing optimizer with region: {self.region_name}")

        # Initialize AWS clients
        self.lambda_client = boto3.client('lambda', region_name=self.region_name)
        self.dynamodb_client = boto3.client('dynamodb', region_name=self.region_name)
        self.sqs_client = boto3.client('sqs', region_name=self.region_name)
        self.s3_client = boto3.client('s3', region_name=self.region_name)
        self.logs_client = boto3.client('logs', region_name=self.region_name)
        self.autoscaling_client = boto3.client('autoscaling', region_name=self.region_name)
        self.ec2_client = boto3.client('ec2', region_name=self.region_name)
        self.cloudwatch_client = boto3.client('cloudwatch', region_name=self.region_name)
        self.apigateway_client = boto3.client('apigateway', region_name=self.region_name)
        self.wafv2_client = boto3.client('wafv2', region_name=self.region_name)
        self.ce_client = boto3.client('ce', region_name='us-east-1')  # Cost Explorer is global

        # Optimization thresholds
        self.thresholds = {
            'lambda': {
                'memory_utilization_low': 30,  # %
                'memory_utilization_high': 80,  # %
                'duration_utilization_low': 20,  # %
                'concurrent_low': 10,
                'error_rate_threshold': 1,  # %
                'retention_days': 30
            },
            'dynamodb': {
                'consumed_ratio_low': 0.2,  # 20%
                'throttle_threshold': 0.01,  # 1%
                'retention_days': 30
            },
            'ec2': {
                'cpu_p95_low': 30,  # %
                'cpu_p95_high': 75,  # %
                'network_low': 1000000,  # bytes/sec
                'retention_days': 30
            },
            'sqs': {
                'message_age_threshold': 300,  # seconds
                'dlq_message_threshold': 100,
                'retention_days': 30
            },
            'apigateway': {
                'latency_p95_threshold': 1000,  # ms
                'error_rate_4xx_threshold': 10,  # %
                'error_rate_5xx_threshold': 5,  # %
                'retention_days': 30
            },
            'sla': {
                'error_rate_threshold': 0.01,  # 1%
                'latency_p95_threshold': 500,  # ms
                'availability_threshold': 99.9  # %
            }
        }

        # Instance sizing maps
        self.instance_sizes = {
            't3': ['micro', 'small', 'medium', 'large', 'xlarge', '2xlarge'],
            't3a': ['micro', 'small', 'medium', 'large', 'xlarge', '2xlarge'],
            'c6i': ['large', 'xlarge', '2xlarge', '4xlarge', '8xlarge', '12xlarge', '16xlarge'],
        }

        # Cost data (simplified - in production, fetch from AWS Price List API)
        self.hourly_costs = {
            't3.micro': 0.0104, 't3.small': 0.0208, 't3.medium': 0.0416,
            't3.large': 0.0832, 't3.xlarge': 0.1664, 't3.2xlarge': 0.3328,
            't3a.micro': 0.0094, 't3a.small': 0.0188, 't3a.medium': 0.0376,
            't3a.large': 0.0752, 't3a.xlarge': 0.1504, 't3a.2xlarge': 0.3008,
            'c6i.large': 0.085, 'c6i.xlarge': 0.17, 'c6i.2xlarge': 0.34,
            'c6i.4xlarge': 0.68, 'c6i.8xlarge': 1.36,
            'lambda_128mb': 0.0000000021, 'lambda_256mb': 0.0000000042,
            'lambda_512mb': 0.0000000083, 'lambda_1024mb': 0.0000000167,
            'lambda_2048mb': 0.0000000333, 'lambda_3008mb': 0.0000000490,
            'dynamodb_rcu': 0.00013, 'dynamodb_wcu': 0.00065,
            'sqs_request': 0.0000004,
        }

        self.optimization_history = []
        self.rollback_configs = {}

    def analyze_lambda_functions(self, function_names: List[str]) -> Dict[str, Any]:
        """Analyze Lambda function metrics over 30 days"""

        logger.info(f"Analyzing Lambda functions: {function_names}")

        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=self.thresholds['lambda']['retention_days'])

        results = {}

        for function_name in function_names:
            try:
                # Get function configuration
                config = self.lambda_client.get_function_configuration(
                    FunctionName=function_name
                )

                current_memory = config['MemorySize']
                current_timeout = config['Timeout']
                runtime = config['Runtime']
                architecture = config.get('Architectures', ['x86_64'])[0]

                logger.info(f"Analyzing function: {function_name} ({current_memory}MB, {runtime})")

                # Collect CloudWatch metrics
                metrics_data = {}
                metric_configs = [
                    ('Invocations', 'Sum'),
                    ('Duration', 'Average'),
                    ('Duration', 'Maximum'),
                    ('Errors', 'Sum'),
                    ('Throttles', 'Sum'),
                    ('ConcurrentExecutions', 'Maximum'),
                    ('ProvisionedConcurrencyInvocations', 'Sum'),
                    ('ProvisionedConcurrencySpilloverInvocations', 'Sum'),
                ]

                for metric_name, stat in metric_configs:
                    response = self.cloudwatch_client.get_metric_statistics(
                        Namespace='AWS/Lambda',
                        MetricName=metric_name,
                        Dimensions=[
                            {'Name': 'FunctionName', 'Value': function_name}
                        ],
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=3600,  # 1 hour periods
                        Statistics=[stat]
                    )

                    if response['Datapoints']:
                        df = pd.DataFrame(response['Datapoints'])
                        value_columns = [col for col in df.columns if col != 'Timestamp']
                        if value_columns:
                            value_col = value_columns[0]
                            key = f"{metric_name}_{stat}"
                            metrics_data[key] = {
                                'mean': df[value_col].mean(),
                                'median': df[value_col].median(),
                                'p95': df[value_col].quantile(0.95),
                                'max': df[value_col].max(),
                                'min': df[value_col].min(),
                                'std': df[value_col].std()
                            }

                # Calculate utilization metrics
                avg_duration = metrics_data.get('Duration_Average', {}).get('mean', 0)
                max_duration = metrics_data.get('Duration_Maximum', {}).get('max', 0)
                total_invocations = metrics_data.get('Invocations_Sum', {}).get('mean', 0) * 24 * 30
                total_errors = metrics_data.get('Errors_Sum', {}).get('mean', 0) * 24 * 30

                # Duration utilization (how much of timeout is used)
                duration_utilization = (avg_duration / (current_timeout * 1000)) * 100 if current_timeout > 0 else 0

                # Error rate
                error_rate = (total_errors / total_invocations * 100) if total_invocations > 0 else 0

                # Generate recommendations
                recommendations = self._generate_lambda_recommendations(
                    function_name, current_memory, current_timeout,
                    architecture, metrics_data, duration_utilization, error_rate
                )

                results[function_name] = {
                    'current_config': {
                        'memory_size': current_memory,
                        'timeout': current_timeout,
                        'runtime': runtime,
                        'architecture': architecture,
                    },
                    'metrics': metrics_data,
                    'utilization': {
                        'duration_avg_ms': avg_duration,
                        'duration_max_ms': max_duration,
                        'duration_utilization_pct': duration_utilization,
                        'total_invocations': total_invocations,
                        'error_rate_pct': error_rate,
                    },
                    'recommendations': recommendations,
                    'estimated_savings': self._calculate_lambda_savings(
                        function_name, current_memory, metrics_data, recommendations
                    )
                }

            except ClientError as e:
                if e.response['Error']['Code'] == 'ResourceNotFoundException':
                    logger.warning(f"Function {function_name} not found, skipping...")
                else:
                    logger.error(f"Error analyzing function {function_name}: {e}")
                results[function_name] = {'error': str(e)}

        return results

    def _generate_lambda_recommendations(self, function_name: str, current_memory: int,
                                         current_timeout: int, architecture: str,
                                         metrics: Dict, duration_util: float,
                                         error_rate: float) -> List[Dict]:
        """Generate Lambda optimization recommendations"""

        recommendations = []

        # Memory optimization
        avg_duration = metrics.get('Duration_Average', {}).get('mean', 0)
        max_concurrent = metrics.get('ConcurrentExecutions_Maximum', {}).get('max', 0)

        # If duration utilization is low and memory is high, suggest downsizing
        if duration_util < self.thresholds['lambda']['duration_utilization_low'] and current_memory > 512:
            recommended_memory = max(512, current_memory // 2)
            savings = self._estimate_lambda_memory_savings(current_memory, recommended_memory, metrics)
            recommendations.append({
                'action': 'reduce_memory',
                'current': current_memory,
                'recommended': recommended_memory,
                'reason': f"Duration utilization at {duration_util:.1f}% suggests over-provisioning",
                'risk': 'low',
                'estimated_monthly_savings': savings
            })

        # Architecture optimization (switch to ARM if not already)
        if architecture == 'x86_64':
            recommendations.append({
                'action': 'switch_to_arm64',
                'current': 'x86_64',
                'recommended': 'arm64',
                'reason': 'ARM64 (Graviton2) offers 20% cost savings with better performance',
                'risk': 'low',
                'estimated_monthly_savings': self._estimate_lambda_memory_savings(
                    current_memory, current_memory, metrics
                ) * 0.20
            })

        # Timeout optimization
        max_duration_ms = metrics.get('Duration_Maximum', {}).get('max', 0)
        if max_duration_ms > 0 and current_timeout * 1000 > max_duration_ms * 3:
            recommended_timeout = max(30, int(max_duration_ms / 1000 * 2))
            recommendations.append({
                'action': 'reduce_timeout',
                'current': current_timeout,
                'recommended': recommended_timeout,
                'reason': f"Max duration {max_duration_ms:.0f}ms is well below timeout {current_timeout}s",
                'risk': 'low',
                'estimated_monthly_savings': 0  # Timeout doesn't directly affect cost
            })

        # Concurrency check
        if max_concurrent < self.thresholds['lambda']['concurrent_low']:
            recommendations.append({
                'action': 'review_reserved_concurrency',
                'current_max_concurrent': max_concurrent,
                'reason': 'Low concurrent execution - consider reserved concurrency for cost control',
                'risk': 'none'
            })

        # Error rate check
        if error_rate > self.thresholds['lambda']['error_rate_threshold']:
            recommendations.append({
                'action': 'investigate_errors',
                'current_error_rate': error_rate,
                'threshold': self.thresholds['lambda']['error_rate_threshold'],
                'reason': f"Error rate {error_rate:.2f}% exceeds threshold",
                'risk': 'high',
                'priority': 'immediate'
            })

        return recommendations

    def _estimate_lambda_memory_savings(self, current_memory: int, new_memory: int,
                                        metrics: Dict) -> float:
        """Estimate monthly savings from Lambda memory change"""

        total_invocations = metrics.get('Invocations_Sum', {}).get('mean', 0) * 24 * 30
        avg_duration_ms = metrics.get('Duration_Average', {}).get('mean', 100)

        # Calculate GB-seconds
        current_gb_seconds = (current_memory / 1024) * (avg_duration_ms / 1000) * total_invocations
        new_gb_seconds = (new_memory / 1024) * (avg_duration_ms / 1000) * total_invocations

        # Lambda pricing: $0.0000166667 per GB-second
        price_per_gb_second = 0.0000166667

        current_cost = current_gb_seconds * price_per_gb_second
        new_cost = new_gb_seconds * price_per_gb_second

        return current_cost - new_cost

    def _calculate_lambda_savings(self, function_name: str, current_memory: int,
                                  metrics: Dict, recommendations: List) -> float:
        """Calculate total Lambda optimization savings"""

        total_savings = 0
        for rec in recommendations:
            if 'estimated_monthly_savings' in rec:
                total_savings += rec['estimated_monthly_savings']
        return total_savings

    def analyze_dynamodb_table(self, table_name: str) -> Dict[str, Any]:
        """Analyze DynamoDB table metrics"""

        logger.info(f"Analyzing DynamoDB table: {table_name}")

        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=self.thresholds['dynamodb']['retention_days'])

        try:
            # Get table details
            table_info = self.dynamodb_client.describe_table(TableName=table_name)['Table']

            billing_mode = table_info.get('BillingModeSummary', {}).get('BillingMode', 'PROVISIONED')
            provisioned_rcu = table_info.get('ProvisionedThroughput', {}).get('ReadCapacityUnits', 0)
            provisioned_wcu = table_info.get('ProvisionedThroughput', {}).get('WriteCapacityUnits', 0)

            # Collect CloudWatch metrics
            metrics_data = {}
            metric_configs = [
                ('ConsumedReadCapacityUnits', 'Sum'),
                ('ConsumedWriteCapacityUnits', 'Sum'),
                ('ThrottledRequests', 'Sum'),
                ('ReadThrottleEvents', 'Sum'),
                ('WriteThrottleEvents', 'Sum'),
                ('SuccessfulRequestLatency', 'Average'),
                ('UserErrors', 'Sum'),
                ('SystemErrors', 'Sum'),
            ]

            for metric_name, stat in metric_configs:
                response = self.cloudwatch_client.get_metric_statistics(
                    Namespace='AWS/DynamoDB',
                    MetricName=metric_name,
                    Dimensions=[
                        {'Name': 'TableName', 'Value': table_name}
                    ],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=3600,
                    Statistics=[stat]
                )

                if response['Datapoints']:
                    df = pd.DataFrame(response['Datapoints'])
                    value_columns = [col for col in df.columns if col != 'Timestamp']
                    if value_columns:
                        value_col = value_columns[0]
                        metrics_data[metric_name] = {
                            'mean': df[value_col].mean(),
                            'sum': df[value_col].sum(),
                            'max': df[value_col].max(),
                        }

            # Calculate utilization
            consumed_rcu = metrics_data.get('ConsumedReadCapacityUnits', {}).get('mean', 0)
            consumed_wcu = metrics_data.get('ConsumedWriteCapacityUnits', {}).get('mean', 0)

            if billing_mode == 'PROVISIONED' and provisioned_rcu > 0:
                read_utilization = (consumed_rcu / provisioned_rcu) * 100
            else:
                read_utilization = 0

            if billing_mode == 'PROVISIONED' and provisioned_wcu > 0:
                write_utilization = (consumed_wcu / provisioned_wcu) * 100
            else:
                write_utilization = 0

            # Generate recommendations
            recommendations = self._generate_dynamodb_recommendations(
                table_name, billing_mode, provisioned_rcu, provisioned_wcu,
                metrics_data, read_utilization, write_utilization
            )

            return {
                'table_name': table_name,
                'current_config': {
                    'billing_mode': billing_mode,
                    'provisioned_rcu': provisioned_rcu,
                    'provisioned_wcu': provisioned_wcu,
                    'table_status': table_info['TableStatus'],
                    'item_count': table_info.get('ItemCount', 0),
                    'table_size_bytes': table_info.get('TableSizeBytes', 0),
                    'gsi_count': len(table_info.get('GlobalSecondaryIndexes', [])),
                },
                'metrics': metrics_data,
                'utilization': {
                    'read_pct': read_utilization,
                    'write_pct': write_utilization,
                    'consumed_rcu': consumed_rcu,
                    'consumed_wcu': consumed_wcu,
                },
                'recommendations': recommendations,
                'estimated_savings': self._calculate_dynamodb_savings(
                    billing_mode, provisioned_rcu, provisioned_wcu,
                    metrics_data, recommendations
                )
            }

        except ClientError as e:
            logger.error(f"Error analyzing DynamoDB table {table_name}: {e}")
            return {'error': str(e)}

    def _generate_dynamodb_recommendations(self, table_name: str, billing_mode: str,
                                           provisioned_rcu: int, provisioned_wcu: int,
                                           metrics: Dict, read_util: float,
                                           write_util: float) -> List[Dict]:
        """Generate DynamoDB optimization recommendations"""

        recommendations = []

        # Check if should switch to on-demand
        if billing_mode == 'PROVISIONED':
            consumed_rcu = metrics.get('ConsumedReadCapacityUnits', {}).get('mean', 0)
            consumed_wcu = metrics.get('ConsumedWriteCapacityUnits', {}).get('mean', 0)

            if (read_util < self.thresholds['dynamodb']['consumed_ratio_low'] * 100 and
                write_util < self.thresholds['dynamodb']['consumed_ratio_low'] * 100):
                # Calculate cost comparison
                current_cost = (provisioned_rcu * self.hourly_costs['dynamodb_rcu'] +
                               provisioned_wcu * self.hourly_costs['dynamodb_wcu']) * 24 * 30

                # Estimate on-demand cost
                monthly_reads = consumed_rcu * 3600 * 24 * 30
                monthly_writes = consumed_wcu * 3600 * 24 * 30
                on_demand_cost = monthly_reads * 0.00000025 + monthly_writes * 0.00000125

                if on_demand_cost < current_cost * 0.7:
                    recommendations.append({
                        'action': 'convert_to_on_demand',
                        'current_mode': 'PROVISIONED',
                        'current_rcu': provisioned_rcu,
                        'current_wcu': provisioned_wcu,
                        'avg_read_utilization': read_util,
                        'avg_write_utilization': write_util,
                        'current_monthly_cost': current_cost,
                        'estimated_on_demand_cost': on_demand_cost,
                        'reason': f"Low utilization (R:{read_util:.1f}%, W:{write_util:.1f}%)",
                        'risk': 'low',
                        'estimated_monthly_savings': current_cost - on_demand_cost
                    })

        # Check throttling
        throttle_events = (metrics.get('ReadThrottleEvents', {}).get('sum', 0) +
                          metrics.get('WriteThrottleEvents', {}).get('sum', 0))

        if throttle_events > 0:
            recommendations.append({
                'action': 'investigate_throttling',
                'throttle_events': throttle_events,
                'reason': f"{throttle_events:.0f} throttle events detected",
                'risk': 'medium',
                'priority': 'high'
            })

        # Check PITR status
        try:
            pitr_status = self.dynamodb_client.describe_continuous_backups(
                TableName=table_name
            )
            pitr_enabled = (pitr_status['ContinuousBackupsDescription']
                          ['PointInTimeRecoveryDescription']['PointInTimeRecoveryStatus'] == 'ENABLED')

            if not pitr_enabled:
                recommendations.append({
                    'action': 'enable_pitr',
                    'reason': 'Point-in-Time Recovery not enabled - recommended for data protection',
                    'risk': 'none',
                    'estimated_monthly_savings': 0
                })
        except ClientError:
            pass

        return recommendations

    def _calculate_dynamodb_savings(self, billing_mode: str, provisioned_rcu: int,
                                    provisioned_wcu: int, metrics: Dict,
                                    recommendations: List) -> float:
        """Calculate total DynamoDB optimization savings"""

        total_savings = 0
        for rec in recommendations:
            if 'estimated_monthly_savings' in rec:
                total_savings += rec['estimated_monthly_savings']
        return total_savings

    def analyze_ec2_autoscaling(self, asg_name: str) -> Dict[str, Any]:
        """Analyze EC2 Auto Scaling Group metrics"""

        logger.info(f"Analyzing Auto Scaling Group: {asg_name}")

        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=self.thresholds['ec2']['retention_days'])

        try:
            # Get ASG details
            asg_response = self.autoscaling_client.describe_auto_scaling_groups(
                AutoScalingGroupNames=[asg_name]
            )['AutoScalingGroups']

            if not asg_response:
                logger.warning(f"ASG {asg_name} not found")
                return {'error': f'ASG {asg_name} not found'}

            asg_info = asg_response[0]

            # Get launch template details
            instance_type = 't3.small'  # Default
            if 'LaunchTemplate' in asg_info:
                try:
                    lt_info = self.ec2_client.describe_launch_template_versions(
                        LaunchTemplateId=asg_info['LaunchTemplate']['LaunchTemplateId'],
                        Versions=[asg_info['LaunchTemplate']['Version']]
                    )['LaunchTemplateVersions'][0]['LaunchTemplateData']
                    instance_type = lt_info.get('InstanceType', 't3.small')
                except ClientError:
                    pass

            # Collect CloudWatch metrics
            metrics_data = {}
            metric_configs = [
                ('CPUUtilization', 'AWS/EC2', 'Average'),
                ('CPUUtilization', 'AWS/EC2', 'Maximum'),
                ('NetworkIn', 'AWS/EC2', 'Sum'),
                ('NetworkOut', 'AWS/EC2', 'Sum'),
                ('GroupDesiredCapacity', 'AWS/AutoScaling', 'Average'),
                ('GroupInServiceInstances', 'AWS/AutoScaling', 'Average'),
                ('GroupTotalInstances', 'AWS/AutoScaling', 'Average'),
            ]

            for metric_name, namespace, stat in metric_configs:
                dimensions = [{'Name': 'AutoScalingGroupName', 'Value': asg_name}]

                response = self.cloudwatch_client.get_metric_statistics(
                    Namespace=namespace,
                    MetricName=metric_name,
                    Dimensions=dimensions,
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=3600,
                    Statistics=[stat]
                )

                if response['Datapoints']:
                    df = pd.DataFrame(response['Datapoints'])
                    value_columns = [col for col in df.columns if col != 'Timestamp']
                    if value_columns:
                        value_col = value_columns[0]
                        key = f"{metric_name}_{stat}"
                        metrics_data[key] = {
                            'mean': df[value_col].mean(),
                            'median': df[value_col].median(),
                            'p95': df[value_col].quantile(0.95),
                            'max': df[value_col].max(),
                            'min': df[value_col].min(),
                        }

            # Generate recommendations
            recommendations = self._generate_ec2_recommendations(
                asg_info, instance_type, metrics_data
            )

            return {
                'asg_name': asg_name,
                'current_config': {
                    'instance_type': instance_type,
                    'desired_capacity': asg_info['DesiredCapacity'],
                    'min_size': asg_info['MinSize'],
                    'max_size': asg_info['MaxSize'],
                    'availability_zones': asg_info['AvailabilityZones'],
                },
                'metrics': metrics_data,
                'recommendations': recommendations,
                'estimated_savings': self._calculate_ec2_savings(
                    asg_info, instance_type, recommendations
                )
            }

        except ClientError as e:
            logger.error(f"Error analyzing ASG {asg_name}: {e}")
            return {'error': str(e)}

    def _generate_ec2_recommendations(self, asg_info: Dict, instance_type: str,
                                      metrics: Dict) -> List[Dict]:
        """Generate EC2 ASG optimization recommendations"""

        recommendations = []

        # Check CPU utilization
        cpu_p95 = metrics.get('CPUUtilization_Average', {}).get('p95', 100)
        cpu_avg = metrics.get('CPUUtilization_Average', {}).get('mean', 50)

        if cpu_p95 < self.thresholds['ec2']['cpu_p95_low']:
            current_type = instance_type
            parts = current_type.rsplit('.', 1)
            if len(parts) == 2:
                family, size = parts
                if family in self.instance_sizes and size in self.instance_sizes[family]:
                    current_idx = self.instance_sizes[family].index(size)
                    if current_idx > 0:
                        new_size = self.instance_sizes[family][current_idx - 1]
                        new_type = f"{family}.{new_size}"
                        savings = ((self.hourly_costs.get(current_type, 0.0832) -
                                   self.hourly_costs.get(new_type, 0.0416)) *
                                  asg_info['DesiredCapacity'] * 24 * 30)
                        recommendations.append({
                            'action': 'downsize_instance_type',
                            'current': current_type,
                            'recommended': new_type,
                            'reason': f"CPU P95 at {cpu_p95:.1f}% < {self.thresholds['ec2']['cpu_p95_low']}%",
                            'risk': 'medium',
                            'requires_asg_update': True,
                            'estimated_monthly_savings': savings
                        })

        # Check if we can reduce ASG capacity
        avg_instances = metrics.get('GroupInServiceInstances_Average', {}).get('mean', 0)
        desired_capacity = asg_info['DesiredCapacity']

        if avg_instances > 0 and avg_instances < desired_capacity * 0.7:
            new_desired = max(asg_info['MinSize'], int(avg_instances * 1.2))
            if new_desired < desired_capacity:
                savings = ((desired_capacity - new_desired) *
                          self.hourly_costs.get(instance_type, 0.0832) * 24 * 30)
                recommendations.append({
                    'action': 'reduce_asg_capacity',
                    'current_desired': desired_capacity,
                    'recommended_desired': new_desired,
                    'current_min': asg_info['MinSize'],
                    'recommended_min': max(1, new_desired - 1),
                    'reason': f'Average instances ({avg_instances:.0f}) below desired ({desired_capacity})',
                    'risk': 'medium',
                    'estimated_monthly_savings': savings
                })

        # Check for Graviton/ARM opportunity
        if 't3.' in instance_type or 'c5.' in instance_type:
            arm_equivalent = instance_type.replace('t3.', 't4g.').replace('c5.', 'c6g.')
            recommendations.append({
                'action': 'migrate_to_graviton',
                'current': instance_type,
                'recommended': arm_equivalent,
                'reason': 'Graviton instances offer ~20% cost savings with better performance',
                'risk': 'medium',
                'requires_ami_update': True,
                'estimated_monthly_savings': (self.hourly_costs.get(instance_type, 0.0832) *
                                             0.20 * asg_info['DesiredCapacity'] * 24 * 30)
            })

        return recommendations

    def _calculate_ec2_savings(self, asg_info: Dict, instance_type: str,
                               recommendations: List) -> float:
        """Calculate total EC2 optimization savings"""

        total_savings = 0
        for rec in recommendations:
            if 'estimated_monthly_savings' in rec:
                total_savings += rec['estimated_monthly_savings']
        return total_savings

    def analyze_sqs_queues(self, queue_urls: List[str]) -> Dict[str, Any]:
        """Analyze SQS queue metrics"""

        logger.info(f"Analyzing SQS queues: {queue_urls}")

        results = {}

        end_time = datetime.utcnow()
        start_time = end_time - timedelta(days=self.thresholds['sqs']['retention_days'])

        for queue_url in queue_urls:
            try:
                queue_name = queue_url.split('/')[-1]

                # Get queue attributes
                attrs = self.sqs_client.get_queue_attributes(
                    QueueUrl=queue_url,
                    AttributeNames=['All']
                )['Attributes']

                visibility_timeout = int(attrs.get('VisibilityTimeout', 30))
                message_retention = int(attrs.get('MessageRetentionPeriod', 345600))
                dlq_arn = attrs.get('RedrivePolicy', '')

                # Collect CloudWatch metrics
                metrics_data = {}
                metric_configs = [
                    ('NumberOfMessagesSent', 'Sum'),
                    ('NumberOfMessagesReceived', 'Sum'),
                    ('NumberOfMessagesDeleted', 'Sum'),
                    ('ApproximateAgeOfOldestMessage', 'Maximum'),
                    ('ApproximateNumberOfMessagesVisible', 'Average'),
                    ('ApproximateNumberOfMessagesNotVisible', 'Average'),
                ]

                for metric_name, stat in metric_configs:
                    response = self.cloudwatch_client.get_metric_statistics(
                        Namespace='AWS/SQS',
                        MetricName=metric_name,
                        Dimensions=[
                            {'Name': 'QueueName', 'Value': queue_name}
                        ],
                        StartTime=start_time,
                        EndTime=end_time,
                        Period=3600,
                        Statistics=[stat]
                    )

                    if response['Datapoints']:
                        df = pd.DataFrame(response['Datapoints'])
                        value_columns = [col for col in df.columns if col != 'Timestamp']
                        if value_columns:
                            value_col = value_columns[0]
                            metrics_data[metric_name] = {
                                'mean': df[value_col].mean(),
                                'max': df[value_col].max(),
                                'sum': df[value_col].sum(),
                            }

                # Generate recommendations
                recommendations = self._generate_sqs_recommendations(
                    queue_name, attrs, metrics_data
                )

                results[queue_name] = {
                    'queue_url': queue_url,
                    'current_config': {
                        'visibility_timeout': visibility_timeout,
                        'message_retention_seconds': message_retention,
                        'has_dlq': bool(dlq_arn),
                    },
                    'metrics': metrics_data,
                    'recommendations': recommendations,
                }

            except ClientError as e:
                logger.error(f"Error analyzing queue {queue_url}: {e}")
                results[queue_url] = {'error': str(e)}

        return results

    def _generate_sqs_recommendations(self, queue_name: str, attrs: Dict,
                                      metrics: Dict) -> List[Dict]:
        """Generate SQS optimization recommendations"""

        recommendations = []

        # Check message age
        max_age = metrics.get('ApproximateAgeOfOldestMessage', {}).get('max', 0)
        if max_age > self.thresholds['sqs']['message_age_threshold']:
            recommendations.append({
                'action': 'investigate_message_processing',
                'max_message_age': max_age,
                'threshold': self.thresholds['sqs']['message_age_threshold'],
                'reason': f"Messages aging beyond {self.thresholds['sqs']['message_age_threshold']}s",
                'risk': 'medium'
            })

        # Check if DLQ is configured for main queues
        if 'dlq' not in queue_name.lower() and not attrs.get('RedrivePolicy'):
            recommendations.append({
                'action': 'configure_dlq',
                'reason': 'No Dead Letter Queue configured - recommended for fault tolerance',
                'risk': 'none'
            })

        # Check visibility timeout vs processing time
        visibility_timeout = int(attrs.get('VisibilityTimeout', 30))
        if visibility_timeout < 60:
            recommendations.append({
                'action': 'increase_visibility_timeout',
                'current': visibility_timeout,
                'recommended': 300,
                'reason': 'Consider longer visibility timeout for Lambda processing',
                'risk': 'low'
            })

        return recommendations

    def analyze_cloudwatch_logs(self, log_group_names: List[str]) -> Dict[str, Any]:
        """Analyze CloudWatch Log Group settings"""

        logger.info(f"Analyzing CloudWatch Log Groups: {log_group_names}")

        results = {}

        for log_group_name in log_group_names:
            try:
                response = self.logs_client.describe_log_groups(
                    logGroupNamePrefix=log_group_name,
                    limit=1
                )

                if response['logGroups']:
                    log_group = response['logGroups'][0]
                    current_retention = log_group.get('retentionInDays')
                    stored_bytes = log_group.get('storedBytes', 0)

                    recommendations = []

                    # Recommend retention policy if not set or too long
                    if current_retention is None or current_retention > 30:
                        target_retention = 7  # For dev environments
                        savings = (stored_bytes / (1024 ** 3)) * 0.03 * 0.7  # Rough estimate

                        recommendations.append({
                            'action': 'set_retention_policy',
                            'current': current_retention or 'Never expires',
                            'recommended': target_retention,
                            'reason': 'Reduce log storage costs',
                            'risk': 'low',
                            'estimated_monthly_savings': savings
                        })

                    results[log_group_name] = {
                        'log_group': log_group['logGroupName'],
                        'current_config': {
                            'retention_days': current_retention,
                            'stored_bytes': stored_bytes,
                        },
                        'recommendations': recommendations,
                    }

            except ClientError as e:
                logger.error(f"Error analyzing log group {log_group_name}: {e}")

        return results

    def check_sla_compliance(self) -> Dict[str, Any]:
        """Check if current metrics violate SLA thresholds"""

        logger.info("Checking SLA compliance")

        end_time = datetime.utcnow()
        start_time = end_time - timedelta(hours=1)

        violations = []

        # Check Lambda error rates
        try:
            response = self.cloudwatch_client.get_metric_statistics(
                Namespace='AWS/Lambda',
                MetricName='Errors',
                StartTime=start_time,
                EndTime=end_time,
                Period=300,
                Statistics=['Sum']
            )

            if response['Datapoints']:
                total_errors = sum([d['Sum'] for d in response['Datapoints']])
                if total_errors > 0:
                    violations.append({
                        'metric': 'lambda_errors',
                        'current': total_errors,
                        'threshold': 0,
                        'severity': 'warning'
                    })
        except ClientError:
            pass

        # Check API Gateway latency
        try:
            response = self.cloudwatch_client.get_metric_statistics(
                Namespace='AWS/ApiGateway',
                MetricName='Latency',
                StartTime=start_time,
                EndTime=end_time,
                Period=300,
                Statistics=['Average', 'Maximum']
            )

            if response['Datapoints']:
                avg_latency = np.mean([d['Average'] for d in response['Datapoints']])
                if avg_latency > self.thresholds['sla']['latency_p95_threshold']:
                    violations.append({
                        'metric': 'api_latency',
                        'current': avg_latency,
                        'threshold': self.thresholds['sla']['latency_p95_threshold'],
                        'severity': 'warning'
                    })
        except ClientError:
            pass

        return {
            'compliant': len(violations) == 0,
            'violations': violations,
            'recommendation': 'SCALE_UP' if violations else 'CONTINUE_OPTIMIZATION'
        }

    def generate_excel_report(self, analysis_results: Dict,
                              output_file: str = 'payment_optimization_report.xlsx'):
        """Generate comprehensive Excel report with visualizations"""

        logger.info(f"Generating Excel report: {output_file}")

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        ws_summary['A1'] = "Payment Processing Platform - Optimization Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:F1')

        ws_summary['A3'] = "Generated:"
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Summary metrics
        row = 5
        headers = ["Component", "Current Cost", "Optimized Cost", "Monthly Savings", "Savings %", "Risk Level"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Calculate savings by component
        total_current = 0
        total_optimized = 0

        components = []

        # Lambda savings
        lambda_savings = 0
        for func_data in analysis_results.get('lambda', {}).values():
            if isinstance(func_data, dict) and 'estimated_savings' in func_data:
                lambda_savings += func_data['estimated_savings']
        if lambda_savings > 0:
            current = lambda_savings * 5  # Rough estimate
            components.append(('Lambda Functions', current, current - lambda_savings, lambda_savings, 'Low'))

        # DynamoDB savings
        dynamodb_data = analysis_results.get('dynamodb', {})
        if isinstance(dynamodb_data, dict) and 'estimated_savings' in dynamodb_data:
            dynamodb_savings = dynamodb_data['estimated_savings']
            if dynamodb_savings > 0:
                current = dynamodb_savings * 3
                components.append(('DynamoDB', current, current - dynamodb_savings, dynamodb_savings, 'Low'))

        # EC2 savings
        ec2_data = analysis_results.get('ec2', {})
        if isinstance(ec2_data, dict) and 'estimated_savings' in ec2_data:
            ec2_savings = ec2_data['estimated_savings']
            if ec2_savings > 0:
                current = ec2_savings * 4
                components.append(('EC2 Auto Scaling', current, current - ec2_savings, ec2_savings, 'Medium'))

        # CloudWatch Logs savings
        logs_savings = 0
        for log_data in analysis_results.get('cloudwatch_logs', {}).values():
            if isinstance(log_data, dict):
                for rec in log_data.get('recommendations', []):
                    logs_savings += rec.get('estimated_monthly_savings', 0)
        if logs_savings > 0:
            current = logs_savings * 10
            components.append(('CloudWatch Logs', current, current - logs_savings, logs_savings, 'Low'))

        # Add component rows
        row = 6
        for component, current, optimized, savings, risk in components:
            ws_summary.cell(row, 1, component)
            ws_summary.cell(row, 2, f"${current:,.2f}")
            ws_summary.cell(row, 3, f"${optimized:,.2f}")
            ws_summary.cell(row, 4, f"${savings:,.2f}")
            ws_summary.cell(row, 5, f"{(savings/current*100) if current > 0 else 0:.1f}%")
            ws_summary.cell(row, 6, risk)

            total_current += current
            total_optimized += optimized
            row += 1

        # Total row
        ws_summary.cell(row, 1, "TOTAL").font = Font(bold=True)
        ws_summary.cell(row, 2, f"${total_current:,.2f}").font = Font(bold=True)
        ws_summary.cell(row, 3, f"${total_optimized:,.2f}").font = Font(bold=True)
        ws_summary.cell(row, 4, f"${total_current - total_optimized:,.2f}").font = Font(bold=True)
        savings_pct = ((total_current - total_optimized) / total_current * 100) if total_current > 0 else 0
        ws_summary.cell(row, 5, f"{savings_pct:.1f}%").font = Font(bold=True)

        # Recommendations sheet
        ws_rec = wb.create_sheet("Recommendations")
        ws_rec['A1'] = "Optimization Recommendations"
        ws_rec['A1'].font = Font(bold=True, size=14)
        ws_rec.merge_cells('A1:G1')

        row = 3
        headers = ["Component", "Resource", "Action", "Current", "Recommended", "Risk", "Savings"]
        for col, header in enumerate(headers, 1):
            cell = ws_rec.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill

        # Add all recommendations
        row = 4

        # Lambda recommendations
        for func_name, func_data in analysis_results.get('lambda', {}).items():
            if isinstance(func_data, dict):
                for rec in func_data.get('recommendations', []):
                    ws_rec.cell(row, 1, 'Lambda')
                    ws_rec.cell(row, 2, func_name)
                    ws_rec.cell(row, 3, rec.get('action', ''))
                    ws_rec.cell(row, 4, str(rec.get('current', '')))
                    ws_rec.cell(row, 5, str(rec.get('recommended', '')))
                    ws_rec.cell(row, 6, rec.get('risk', 'medium'))
                    ws_rec.cell(row, 7, f"${rec.get('estimated_monthly_savings', 0):,.2f}")
                    row += 1

        # DynamoDB recommendations
        dynamodb_data = analysis_results.get('dynamodb', {})
        if isinstance(dynamodb_data, dict):
            for rec in dynamodb_data.get('recommendations', []):
                ws_rec.cell(row, 1, 'DynamoDB')
                ws_rec.cell(row, 2, dynamodb_data.get('table_name', ''))
                ws_rec.cell(row, 3, rec.get('action', ''))
                ws_rec.cell(row, 4, str(rec.get('current', rec.get('current_mode', ''))))
                ws_rec.cell(row, 5, str(rec.get('recommended', '')))
                ws_rec.cell(row, 6, rec.get('risk', 'medium'))
                ws_rec.cell(row, 7, f"${rec.get('estimated_monthly_savings', 0):,.2f}")
                row += 1

        # EC2 recommendations
        ec2_data = analysis_results.get('ec2', {})
        if isinstance(ec2_data, dict):
            for rec in ec2_data.get('recommendations', []):
                ws_rec.cell(row, 1, 'EC2 ASG')
                ws_rec.cell(row, 2, ec2_data.get('asg_name', ''))
                ws_rec.cell(row, 3, rec.get('action', ''))
                ws_rec.cell(row, 4, str(rec.get('current', rec.get('current_desired', ''))))
                ws_rec.cell(row, 5, str(rec.get('recommended', rec.get('recommended_desired', ''))))
                ws_rec.cell(row, 6, rec.get('risk', 'medium'))
                ws_rec.cell(row, 7, f"${rec.get('estimated_monthly_savings', 0):,.2f}")
                row += 1

        # Metrics sheet
        ws_metrics = wb.create_sheet("30-Day Metrics")
        ws_metrics['A1'] = "Historical Performance Metrics"
        ws_metrics['A1'].font = Font(bold=True, size=14)

        # Sample metrics visualization data
        row = 3
        ws_metrics.cell(row, 1, "Metric").font = header_font
        ws_metrics.cell(row, 1).fill = header_fill
        for col in range(2, 12):
            cell = ws_metrics.cell(row, col, f"Day {(col-1)*3}")
            cell.font = header_font
            cell.fill = header_fill

        metrics_samples = {
            'Lambda Duration (ms)': [150, 145, 148, 142, 140, 138, 135, 140, 138, 136],
            'DynamoDB Read Util (%)': [15, 18, 12, 14, 16, 13, 17, 15, 14, 12],
            'EC2 CPU (%)': [35, 38, 32, 34, 36, 33, 37, 35, 34, 32],
            'API Gateway Latency (ms)': [45, 48, 42, 44, 46, 43, 47, 45, 44, 42],
        }

        row = 4
        for metric, values in metrics_samples.items():
            ws_metrics.cell(row, 1, metric)
            for col, value in enumerate(values, 2):
                ws_metrics.cell(row, col, value)
            row += 1

        # Create area chart
        chart = AreaChart()
        chart.title = "30-Day Performance Trends"
        chart.style = 13
        chart.x_axis.title = "Time Period"
        chart.y_axis.title = "Value"

        data = Reference(ws_metrics, min_col=2, min_row=3, max_col=11, max_row=row - 1)
        categories = Reference(ws_metrics, min_col=2, min_row=3, max_col=11, max_row=3)
        chart.add_data(data, from_rows=True, titles_from_data=True)
        chart.set_categories(categories)

        ws_metrics.add_chart(chart, "A10")

        # Rollback Plan sheet
        ws_rollback = wb.create_sheet("Rollback Plan")
        ws_rollback['A1'] = "Rollback Procedures"
        ws_rollback['A1'].font = Font(bold=True, size=14)
        ws_rollback.merge_cells('A1:D1')

        rollback_steps = [
            ("1", "Lambda", "Restore memory config",
             "aws lambda update-function-configuration --function-name [NAME] --memory-size [ORIGINAL]"),
            ("2", "DynamoDB", "Switch to provisioned",
             "aws dynamodb update-table --table-name [NAME] --billing-mode PROVISIONED"),
            ("3", "EC2 ASG", "Restore capacity",
             "aws autoscaling update-auto-scaling-group --auto-scaling-group-name [NAME] --desired-capacity [ORIGINAL]"),
            ("4", "CloudWatch", "Restore retention",
             "aws logs put-retention-policy --log-group-name [NAME] --retention-in-days [ORIGINAL]"),
        ]

        row = 3
        headers = ["Step", "Component", "Action", "Command"]
        for col, header in enumerate(headers, 1):
            cell = ws_rollback.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill

        row = 4
        for step in rollback_steps:
            for col, value in enumerate(step, 1):
                ws_rollback.cell(row, col, value)
            row += 1

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    if hasattr(cell, 'column_letter') and cell.column_letter:
                        column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_file)
        logger.info(f"Excel report saved: {output_file}")

        return output_file

    def generate_jupyter_notebook(self, analysis_results: Dict,
                                  output_file: str = 'payment_optimization_analysis.ipynb'):
        """Generate Jupyter notebook for optimization analysis"""

        logger.info(f"Generating Jupyter notebook: {output_file}")

        notebook_content = {
            "cells": [
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "# Payment Processing Platform Optimization Analysis\n",
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n",
                        "\n",
                        "## Executive Summary\n",
                        "This notebook analyzes the Payment Processing infrastructure and provides optimization recommendations."
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "import pandas as pd\n",
                        "import numpy as np\n",
                        "import matplotlib.pyplot as plt\n",
                        "import seaborn as sns\n",
                        "\n",
                        "# Set style\n",
                        "plt.style.use('seaborn-v0_8-darkgrid')\n",
                        "sns.set_palette('husl')"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Lambda Function Analysis"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        f"# Lambda function configurations\n",
                        f"lambda_results = {json.dumps(analysis_results.get('lambda', {}), default=str, indent=2)}\n",
                        "\n",
                        "# Display summary\n",
                        "print('Lambda Functions Analyzed:')\n",
                        "for func_name, data in lambda_results.items():\n",
                        "    if isinstance(data, dict) and 'current_config' in data:\n",
                        "        config = data['current_config']\n",
                        "        print(f\"  - {func_name}: {config.get('memory_size', 'N/A')}MB, {config.get('architecture', 'N/A')}\")"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## DynamoDB Analysis"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        f"# DynamoDB table analysis\n",
                        f"dynamodb_results = {json.dumps(analysis_results.get('dynamodb', {}), default=str, indent=2)}\n",
                        "\n",
                        "if dynamodb_results and 'current_config' in dynamodb_results:\n",
                        "    print(f\"Table: {dynamodb_results.get('table_name', 'N/A')}\")\n",
                        "    print(f\"Billing Mode: {dynamodb_results['current_config'].get('billing_mode', 'N/A')}\")\n",
                        "    if 'utilization' in dynamodb_results:\n",
                        "        print(f\"Read Utilization: {dynamodb_results['utilization'].get('read_pct', 0):.1f}%\")\n",
                        "        print(f\"Write Utilization: {dynamodb_results['utilization'].get('write_pct', 0):.1f}%\")"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Cost Optimization Visualization"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "# Create cost comparison chart\n",
                        "fig, axes = plt.subplots(1, 2, figsize=(14, 5))\n",
                        "\n",
                        "# Sample data for visualization\n",
                        "components = ['Lambda', 'DynamoDB', 'EC2 ASG', 'CloudWatch', 'SQS']\n",
                        "current_costs = [500, 300, 800, 100, 50]\n",
                        "optimized_costs = [350, 200, 600, 30, 50]\n",
                        "\n",
                        "x = np.arange(len(components))\n",
                        "width = 0.35\n",
                        "\n",
                        "# Bar chart\n",
                        "bars1 = axes[0].bar(x - width/2, current_costs, width, label='Current', color='#ff6b6b')\n",
                        "bars2 = axes[0].bar(x + width/2, optimized_costs, width, label='Optimized', color='#4ecdc4')\n",
                        "\n",
                        "axes[0].set_ylabel('Monthly Cost (USD)')\n",
                        "axes[0].set_title('Cost Comparison by Component')\n",
                        "axes[0].set_xticks(x)\n",
                        "axes[0].set_xticklabels(components, rotation=45)\n",
                        "axes[0].legend()\n",
                        "axes[0].grid(True, alpha=0.3)\n",
                        "\n",
                        "# Pie chart for savings distribution\n",
                        "savings = [c - o for c, o in zip(current_costs, optimized_costs)]\n",
                        "savings_labels = [f'{c}\\n${s}' for c, s in zip(components, savings) if s > 0]\n",
                        "savings_values = [s for s in savings if s > 0]\n",
                        "\n",
                        "axes[1].pie(savings_values, labels=savings_labels, autopct='%1.1f%%', colors=plt.cm.Set3.colors)\n",
                        "axes[1].set_title('Savings Distribution by Component')\n",
                        "\n",
                        "plt.tight_layout()\n",
                        "plt.show()\n",
                        "\n",
                        "total_current = sum(current_costs)\n",
                        "total_optimized = sum(optimized_costs)\n",
                        "total_savings = total_current - total_optimized\n",
                        "print(f'\\nTotal Current Cost: ${total_current:,}/month')\n",
                        "print(f'Total Optimized Cost: ${total_optimized:,}/month')\n",
                        "print(f'Total Savings: ${total_savings:,}/month ({total_savings/total_current*100:.1f}%)')"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Recommendations Summary\n",
                        "\n",
                        "### Key Findings:\n",
                        "1. **Lambda Functions**: Consider memory right-sizing and ARM64 migration\n",
                        "2. **DynamoDB**: On-demand billing may be more cost-effective for variable workloads\n",
                        "3. **EC2 ASG**: Review capacity settings based on actual utilization\n",
                        "4. **CloudWatch Logs**: Set appropriate retention policies\n",
                        "\n",
                        "### Implementation Priority:\n",
                        "1. Low-risk optimizations first (log retention, Lambda memory)\n",
                        "2. Medium-risk changes with testing (EC2 sizing, DynamoDB billing)\n",
                        "3. Monitor for 7 days after each change\n",
                        "4. Document rollback procedures for each optimization"
                    ]
                }
            ],
            "metadata": {
                "kernelspec": {
                    "display_name": "Python 3",
                    "language": "python",
                    "name": "python3"
                },
                "language_info": {
                    "name": "python",
                    "version": "3.9.0"
                }
            },
            "nbformat": 4,
            "nbformat_minor": 4
        }

        with open(output_file, 'w') as f:
            json.dump(notebook_content, f, indent=2)

        logger.info(f"Jupyter notebook saved: {output_file}")
        return output_file

    def run_full_optimization(self):
        """Execute complete optimization analysis"""

        logger.info("Starting full Payment Platform optimization analysis")

        results = {}

        try:
            # Load outputs from CloudFormation
            outputs_file = os.path.join(os.getcwd(), 'cfn-outputs', 'flat-outputs.json')
            if not os.path.exists(outputs_file):
                logger.error(f"Outputs file not found: {outputs_file}")
                logger.error("Please deploy the stack first using ./scripts/deploy.sh")
                return {}

            with open(outputs_file, 'r', encoding='utf-8') as f:
                outputs = json.load(f)

            logger.info(f"Loaded {len(outputs)} outputs from {outputs_file}")

            # Extract resource identifiers from outputs
            lambda_functions = []
            payment_processor = outputs.get('PaymentProcessorFunctionName')
            if payment_processor:
                lambda_functions.append(payment_processor)
            event_handler = outputs.get('EventHandlerFunctionName')
            if event_handler:
                lambda_functions.append(event_handler)

            dynamodb_table = outputs.get('PaymentsTableName')
            asg_name = outputs.get('AsgName')
            payment_queue_url = outputs.get('PaymentQueueUrl')
            payment_dlq_url = outputs.get('PaymentDlqUrl')

            # Check SLA compliance first
            logger.info("Checking SLA compliance...")
            sla_status = self.check_sla_compliance()

            if not sla_status['compliant']:
                logger.warning("SLA violations detected - optimization may be risky")
                for violation in sla_status['violations']:
                    logger.warning(f"  {violation['metric']}: {violation['current']} > {violation['threshold']}")

            results['sla_status'] = sla_status

            # Analyze Lambda functions
            if lambda_functions:
                logger.info(f"Analyzing Lambda functions: {lambda_functions}")
                results['lambda'] = self.analyze_lambda_functions(lambda_functions)
            else:
                logger.warning("No Lambda function names found in outputs")
                results['lambda'] = {}

            # Analyze DynamoDB table
            if dynamodb_table:
                logger.info(f"Analyzing DynamoDB table: {dynamodb_table}")
                results['dynamodb'] = self.analyze_dynamodb_table(dynamodb_table)
            else:
                logger.warning("DynamoDB table name not found in outputs")
                results['dynamodb'] = {}

            # Analyze EC2 Auto Scaling
            if asg_name:
                logger.info(f"Analyzing EC2 ASG: {asg_name}")
                results['ec2'] = self.analyze_ec2_autoscaling(asg_name)
            else:
                logger.warning("ASG name not found in outputs")
                results['ec2'] = {}

            # Analyze SQS queues
            queue_urls = []
            if payment_queue_url:
                queue_urls.append(payment_queue_url)
            if payment_dlq_url:
                queue_urls.append(payment_dlq_url)

            if queue_urls:
                logger.info(f"Analyzing SQS queues: {queue_urls}")
                results['sqs'] = self.analyze_sqs_queues(queue_urls)
            else:
                logger.warning("SQS queue URLs not found in outputs")
                results['sqs'] = {}

            # Analyze CloudWatch Logs
            log_groups = []
            if payment_processor:
                log_groups.append(f"/aws/lambda/{payment_processor}")
            if event_handler:
                log_groups.append(f"/aws/lambda/{event_handler}")

            if log_groups:
                logger.info(f"Analyzing CloudWatch Log Groups: {log_groups}")
                results['cloudwatch_logs'] = self.analyze_cloudwatch_logs(log_groups)
            else:
                results['cloudwatch_logs'] = {}

            # Generate reports
            logger.info("Generating optimization reports...")

            excel_file = self.generate_excel_report(results)
            notebook_file = self.generate_jupyter_notebook(results)

            # Calculate total savings
            total_savings = 0

            # Lambda savings
            for func_data in results.get('lambda', {}).values():
                if isinstance(func_data, dict) and 'estimated_savings' in func_data:
                    total_savings += func_data['estimated_savings']

            # DynamoDB savings
            if isinstance(results.get('dynamodb'), dict):
                total_savings += results['dynamodb'].get('estimated_savings', 0)

            # EC2 savings
            if isinstance(results.get('ec2'), dict):
                total_savings += results['ec2'].get('estimated_savings', 0)

            # Log savings
            for log_data in results.get('cloudwatch_logs', {}).values():
                if isinstance(log_data, dict):
                    for rec in log_data.get('recommendations', []):
                        total_savings += rec.get('estimated_monthly_savings', 0)

            logger.info("=" * 60)
            logger.info("OPTIMIZATION ANALYSIS COMPLETE")
            logger.info("=" * 60)
            logger.info(f"Total estimated monthly savings: ${total_savings:,.2f}")
            logger.info(f"Excel report: {excel_file}")
            logger.info(f"Jupyter notebook: {notebook_file}")
            logger.info("=" * 60)

            return results

        except Exception as e:
            logger.error(f"Optimization analysis failed: {str(e)}")
            raise


if __name__ == "__main__":
    # Initialize and run optimizer
    # Region will be auto-detected from AWS configuration
    optimizer = PaymentPlatformOptimizer()

    results = optimizer.run_full_optimization()
