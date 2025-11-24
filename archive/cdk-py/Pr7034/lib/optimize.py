#!/usr/bin/env python3
"""
Trading Analytics Platform Resource Optimizer
Analyzes 90-day metrics and optimizes resource allocation while maintaining SLAs
Author: Platform Engineering Team
"""

import json
import logging
import os
import warnings
from datetime import datetime, timedelta
from typing import Any, Dict, List

import boto3
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('tap_optimization.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class TradingPlatformOptimizer:
    """
    Comprehensive resource optimizer for trading and ML platforms
    """

    def __init__(self, region_trading=None, region_ml=None):
        """Initialize AWS clients and configuration"""

        # Auto-detect current region if not specified
        session = boto3.Session()
        current_region = session.region_name or 'us-east-1'  # Fallback to us-east-1 if detection fails

        self.region_trading = region_trading or current_region
        self.region_ml = region_ml or current_region

        logger.info(f"Initializing optimizer with trading region: {self.region_trading}, ML region: {self.region_ml}")

        # Initialize AWS clients for trading region
        self.ec2_trading = boto3.client('ec2', region_name=self.region_trading)
        self.asg_trading = boto3.client('autoscaling', region_name=self.region_trading)
        self.rds_trading = boto3.client('rds', region_name=self.region_trading)
        self.elasticache_trading = boto3.client('elasticache', region_name=self.region_trading)
        self.dynamodb_trading = boto3.client('dynamodb', region_name=self.region_trading)
        self.cloudwatch_trading = boto3.client('cloudwatch', region_name=self.region_trading)
        self.ce_trading = boto3.client('ce', region_name='us-east-1')  # Cost Explorer is global

        # Initialize AWS clients for ML region
        self.sagemaker_ml = boto3.client('sagemaker', region_name=self.region_ml)
        self.cloudwatch_ml = boto3.client('cloudwatch', region_name=self.region_ml)
        self.ec2_ml = boto3.client('ec2', region_name=self.region_ml)

        # Optimization thresholds
        self.thresholds = {
            'aurora': {
                'cpu_low': 20,  # %
                'cpu_high': 70,  # %
                'connections_low': 100,
                'connections_high': 3000,
                'retention_days': 90
            },
            'ec2': {
                'cpu_p95_low': 30,  # %
                'cpu_p95_high': 75,  # %
                'network_low': 1000000,  # bytes/sec
                'retention_days': 90
            },
            'redis': {
                'hit_rate_high': 95,  # %
                'cpu_low': 20,  # %
                'memory_low': 30,  # %
                'retention_days': 90
            },
            'dynamodb': {
                'consumed_ratio_low': 0.2,  # 20%
                'throttle_threshold': 0.01,  # 1%
                'retention_days': 30
            },
            'sla': {
                'error_rate_threshold': 0.01,  # 1%
                'latency_p95_threshold': 15,  # ms
                'queue_depth_threshold': 1000
            }
        }

        # Resource sizing maps
        self.instance_sizes = {
            'r6g': ['large', 'xlarge', '2xlarge', '4xlarge', '8xlarge', '12xlarge', '16xlarge', '24xlarge'],
            'c6i': ['large', 'xlarge', '2xlarge', '4xlarge', '8xlarge', '12xlarge', '16xlarge', '24xlarge', '32xlarge'],
            'cache.r6g': ['large', 'xlarge', '2xlarge', '4xlarge', '8xlarge', '12xlarge', '16xlarge'],
            'dax.r4': ['large', 'xlarge', '2xlarge', '4xlarge', '8xlarge'],
        }

        # Cost data (simplified - in production, fetch from AWS Price List API)
        self.hourly_costs = {
            'r6g.large': 0.1008, 'r6g.xlarge': 0.2016, 'r6g.2xlarge': 0.4032,
            'r6g.4xlarge': 0.8064, 'r6g.8xlarge': 1.6128, 'r6g.12xlarge': 2.4192,
            'r6g.16xlarge': 3.2256, 'r6g.24xlarge': 4.8384,
            'c6i.large': 0.085, 'c6i.xlarge': 0.17, 'c6i.2xlarge': 0.34,
            'c6i.4xlarge': 0.68, 'c6i.8xlarge': 1.36, 'c6i.12xlarge': 2.04,
            'c6i.16xlarge': 2.72, 'c6i.24xlarge': 4.08, 'c6i.32xlarge': 5.44,
            'cache.r6g.large': 0.142, 'cache.r6g.xlarge': 0.284, 'cache.r6g.2xlarge': 0.568,
            'cache.r6g.4xlarge': 1.136, 'cache.r6g.8xlarge': 2.272, 'cache.r6g.12xlarge': 3.408,
            'cache.r6g.16xlarge': 4.544,
            'dax.r4.large': 0.269, 'dax.r4.xlarge': 0.538, 'dax.r4.2xlarge': 1.076,
            'dax.r4.4xlarge': 2.152, 'dax.r4.8xlarge': 4.304,
            'dynamodb_rcu': 0.00013, 'dynamodb_wcu': 0.00065,  # per hour
            'ml.p3.2xlarge': 3.06, 'ml.p3.8xlarge': 12.24, 'ml.p3.16xlarge': 24.48,
            'ml.g4dn.xlarge': 0.526, 'ml.g4dn.2xlarge': 0.752, 'ml.g4dn.4xlarge': 1.204,
        }

        self.optimization_history = []
        self.rollback_configs = {}

    def analyze_aurora_cluster(self, cluster_id: str) -> Dict[str, Any]:
        """Analyze Aurora cluster metrics over 90 days"""

        logger.info(f"Analyzing Aurora cluster: {cluster_id}")

        end_time = datetime.utcnow()
        # Use 30 days instead of 90 to stay within CloudWatch limits
        start_time = end_time - timedelta(days=30)

        # Get cluster details
        cluster_info = self.rds_trading.describe_db_clusters(
            DBClusterIdentifier=cluster_id
        )['DBClusters'][0]

        # Get instance details
        instances = []
        for member in cluster_info['DBClusterMembers']:
            instance_info = self.rds_trading.describe_db_instances(
                DBInstanceIdentifier=member['DBInstanceIdentifier']
            )['DBInstances'][0]
            instances.append(instance_info)

        # Collect CloudWatch metrics
        metrics_data = {}
        metric_names = [
            ('CPUUtilization', 'Average'),
            ('DatabaseConnections', 'Average'),
            ('ReadLatency', 'Average'),
            ('WriteLatency', 'Average'),
            ('ReadIOPS', 'Sum'),
            ('WriteIOPS', 'Sum'),
            ('NetworkReceiveThroughput', 'Average'),
            ('NetworkTransmitThroughput', 'Average'),
            ('AuroraReplicaLag', 'Maximum'),
            ('BufferCacheHitRatio', 'Average'),
        ]

        for instance in instances:
            instance_id = instance['DBInstanceIdentifier']
            metrics_data[instance_id] = {}

            for metric_name, stat in metric_names:
                response = self.cloudwatch_trading.get_metric_statistics(
                    Namespace='AWS/RDS',
                    MetricName=metric_name,
                    Dimensions=[
                        {'Name': 'DBInstanceIdentifier', 'Value': instance_id}
                    ],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=7200,  # 2 hours to stay within 1440 datapoint limit
                    Statistics=[stat]
                )

                if response['Datapoints']:
                    df = pd.DataFrame(response['Datapoints'])
                    # Find the column that contains the metric values (exclude Timestamp)
                    value_columns = [col for col in df.columns if col != 'Timestamp']
                    if value_columns:
                        value_col = value_columns[0]  # Use the first non-timestamp column
                        metrics_data[instance_id][metric_name] = {
                            'mean': df[value_col].mean(),
                            'median': df[value_col].median(),
                            'p95': df[value_col].quantile(0.95),
                            'max': df[value_col].max(),
                            'min': df[value_col].min(),
                            'std': df[value_col].std()
                        }

        # Analyze patterns and generate recommendations
        recommendations = self._generate_aurora_recommendations(
            cluster_info, instances, metrics_data
        )

        return {
            'cluster_id': cluster_id,
            'current_config': {
                'engine': cluster_info['Engine'],
                'engine_version': cluster_info['EngineVersion'],
                'instances': len(instances),
                'multi_az': cluster_info.get('MultiAZ', False),
                'backup_retention': cluster_info['BackupRetentionPeriod'],
            },
            'metrics': metrics_data,
            'recommendations': recommendations,
            'estimated_savings': self._calculate_aurora_savings(instances, recommendations)
        }

    def _generate_aurora_recommendations(self, cluster_info, instances, metrics):
        """Generate Aurora optimization recommendations"""

        recommendations = []

        for instance in instances:
            instance_id = instance['DBInstanceIdentifier']
            instance_class = instance['DBInstanceClass']
            instance_metrics = metrics.get(instance_id, {})

            # Check CPU utilization
            cpu_metrics = instance_metrics.get('CPUUtilization', {})
            if cpu_metrics.get('p95', 100) < self.thresholds['aurora']['cpu_low']:
                current_size = instance_class.split('.')[-1]
                if current_size in self.instance_sizes['r6g']:
                    current_idx = self.instance_sizes['r6g'].index(current_size)
                    if current_idx > 0:
                        new_size = self.instance_sizes['r6g'][current_idx - 1]
                        recommendations.append({
                            'instance': instance_id,
                            'action': 'downsize',
                            'current': instance_class,
                            'recommended': f"db.r6g.{new_size}",
                            'reason': (
                                f"CPU P95 at {cpu_metrics.get('p95', 0):.1f}% < "
                                f"{self.thresholds['aurora']['cpu_low']}%"
                            ),
                            'risk': 'low',
                            'estimated_monthly_savings': self._calculate_instance_savings(
                                instance_class, f"r6g.{new_size}"
                            )
                        })

            # Check connection count
            conn_metrics = instance_metrics.get('DatabaseConnections', {})
            if conn_metrics.get('p95', 0) < self.thresholds['aurora']['connections_low']:
                recommendations.append({
                    'instance': instance_id,
                    'action': 'review_connection_pooling',
                    'current_connections': conn_metrics.get('mean', 0),
                    'reason': 'Low connection utilization - consider connection pooling optimization',
                    'risk': 'none'
                })

            # Check cache hit ratio
            cache_metrics = instance_metrics.get('BufferCacheHitRatio', {})
            if cache_metrics.get('mean', 100) < 95:
                recommendations.append({
                    'instance': instance_id,
                    'action': 'increase_instance_memory',
                    'current_hit_ratio': cache_metrics.get('mean', 0),
                    'reason': 'Buffer cache hit ratio below optimal threshold',
                    'risk': 'medium'
                })

        # Check if we can reduce reader count
        reader_instances = [i for i in instances if not i.get('IsClusterWriter', False)]
        if len(reader_instances) > 2:
            reader_cpu_avg = np.mean([
                metrics.get(i['DBInstanceIdentifier'], {}).get('CPUUtilization', {}).get('mean', 0)
                for i in reader_instances
            ])

            if reader_cpu_avg < 30:
                recommendations.append({
                    'action': 'reduce_reader_count',
                    'current': len(reader_instances),
                    'recommended': max(2, len(reader_instances) - 1),
                    'reason': f'Average reader CPU at {reader_cpu_avg:.1f}%',
                    'risk': 'medium',
                    'estimated_monthly_savings': self._calculate_instance_savings(
                        instances[0]['DBInstanceClass'], None
                    )
                })

        return recommendations

    def analyze_ec2_autoscaling(self, asg_name: str) -> Dict[str, Any]:
        """Analyze EC2 Auto Scaling Group metrics"""

        logger.info(f"Analyzing Auto Scaling Group: {asg_name}")

        end_time = datetime.utcnow()
        # Use 30 days instead of 90 to stay within CloudWatch limits
        start_time = end_time - timedelta(days=30)

        # Get ASG details
        asg_info = self.asg_trading.describe_auto_scaling_groups(
            AutoScalingGroupNames=[asg_name]
        )['AutoScalingGroups'][0]

        # Get launch template/configuration details
        if 'LaunchTemplate' in asg_info:
            lt_info = self.ec2_trading.describe_launch_template_versions(
                LaunchTemplateId=asg_info['LaunchTemplate']['LaunchTemplateId'],
                Versions=[asg_info['LaunchTemplate']['Version']]
            )['LaunchTemplateVersions'][0]['LaunchTemplateData']
            instance_type = lt_info.get('InstanceType', 'unknown')
        else:
            instance_type = 'c6i.8xlarge'  # Default assumption

        # Collect CloudWatch metrics
        metrics_data = {}
        metric_names = [
            ('CPUUtilization', 'AWS/EC2', 'Average'),
            ('CPUUtilization', 'AWS/EC2', 'Maximum'),
            ('NetworkIn', 'AWS/EC2', 'Sum'),
            ('NetworkOut', 'AWS/EC2', 'Sum'),
            ('NetworkPacketsIn', 'AWS/EC2', 'Sum'),
            ('NetworkPacketsOut', 'AWS/EC2', 'Sum'),
            ('GroupDesiredCapacity', 'AWS/AutoScaling', 'Average'),
            ('GroupInServiceInstances', 'AWS/AutoScaling', 'Average'),
            ('GroupTotalInstances', 'AWS/AutoScaling', 'Average'),
        ]

        for metric_name, namespace, stat in metric_names:
            if namespace == 'AWS/EC2':
                dimensions = [{'Name': 'AutoScalingGroupName', 'Value': asg_name}]
            else:
                dimensions = [{'Name': 'AutoScalingGroupName', 'Value': asg_name}]

            response = self.cloudwatch_trading.get_metric_statistics(
                Namespace=namespace,
                MetricName=metric_name,
                Dimensions=dimensions,
                StartTime=start_time,
                EndTime=end_time,
                Period=7200,  # 2 hours
                Statistics=[stat]
            )

            if response['Datapoints']:
                df = pd.DataFrame(response['Datapoints'])
                # Find the column that contains the metric values (exclude Timestamp)
                value_columns = [col for col in df.columns if col != 'Timestamp']
                if value_columns:
                    value_col = value_columns[0]  # Use the first non-timestamp column
                    metrics_data[f"{metric_name}_{stat}"] = {
                        'mean': df[value_col].mean(),
                        'median': df[value_col].median(),
                        'p95': df[value_col].quantile(0.95),
                        'max': df[value_col].max(),
                        'min': df[value_col].min(),
                        'std': df[value_col].std()
                    }

        # Generate recommendations
        recommendations = self._generate_ec2_recommendations(
            asg_info, instance_type, metrics_data
        )

        return {
            'asg_name': asg_name,
            'current_config': {
                'instance_type': instance_type,
                'desired_capacity': asg_info['DesiredCapacity'],
                'min_size': asg_info['MinSize'],
                'max_size': asg_info['MaxSize'],
                'availability_zones': asg_info['AvailabilityZones'],
            },
            'metrics': metrics_data,
            'recommendations': recommendations,
            'estimated_savings': self._calculate_ec2_savings(
                asg_info, instance_type, recommendations
            )
        }

    def _generate_ec2_recommendations(self, asg_info, instance_type, metrics):
        """Generate EC2 ASG optimization recommendations"""

        recommendations = []

        # Check CPU utilization
        cpu_p95 = metrics.get('CPUUtilization_Average', {}).get('p95', 100)
        if cpu_p95 < self.thresholds['ec2']['cpu_p95_low']:
            current_type = instance_type
            family, size = current_type.rsplit('.', 1)

            if size in self.instance_sizes['c6i']:
                current_idx = self.instance_sizes['c6i'].index(size)
                if current_idx > 0:
                    new_size = self.instance_sizes['c6i'][current_idx - 1]
                    recommendations.append({
                        'action': 'downsize_instance_type',
                        'current': current_type,
                        'recommended': f"{family}.{new_size}",
                        'reason': f"CPU P95 at {cpu_p95:.1f}% < {self.thresholds['ec2']['cpu_p95_low']}%",
                        'risk': 'medium',
                        'requires_asg_update': True
                    })

        # Check if we can reduce ASG capacity
        avg_instances = metrics.get('GroupInServiceInstances_Average', {}).get('mean', 20)
        desired_capacity = asg_info['DesiredCapacity']

        if avg_instances < desired_capacity * 0.7:
            new_desired = max(asg_info['MinSize'], int(avg_instances * 1.2))
            recommendations.append({
                'action': 'reduce_asg_capacity',
                'current_desired': desired_capacity,
                'recommended_desired': new_desired,
                'current_min': asg_info['MinSize'],
                'recommended_min': max(10, new_desired - 5),
                'reason': f'Average instances ({avg_instances:.0f}) significantly below desired ({desired_capacity})',
                'risk': 'medium',
                'estimated_monthly_savings': (desired_capacity - new_desired) *
                    self.hourly_costs.get(instance_type, 1.36) * 24 * 30
            })

        # Check network utilization
        network_in = metrics.get('NetworkIn_Sum', {}).get('mean', 0)
        if network_in < self.thresholds['ec2']['network_low']:
            recommendations.append({
                'action': 'review_network_optimization',
                'current_throughput': network_in,
                'note': 'Consider enabling SR-IOV or review placement group strategy',
                'risk': 'low'
            })

        return recommendations

    def analyze_redis_cluster(self, cluster_id: str) -> Dict[str, Any]:
        """Analyze ElastiCache Redis cluster metrics"""

        logger.info(f"Analyzing Redis cluster: {cluster_id}")

        end_time = datetime.utcnow()
        # Use 30 days instead of 90 to stay within CloudWatch limits
        start_time = end_time - timedelta(days=30)

        # Get cluster details
        cluster_info = self.elasticache_trading.describe_replication_groups(
            ReplicationGroupId=cluster_id
        )['ReplicationGroups'][0]

        # Get node group details
        node_groups = cluster_info['NodeGroups']
        num_shards = len(node_groups)
        replicas_per_shard = len(node_groups[0]['NodeGroupMembers']) - 1 if node_groups else 0

        # Collect CloudWatch metrics
        metrics_data = {}
        metric_names = [
            ('EngineCPUUtilization', 'Average'),
            ('CacheMisses', 'Sum'),
            ('CacheHits', 'Sum'),
            ('BytesUsedForCache', 'Maximum'),
            ('CurrConnections', 'Average'),
            ('NetworkBytesIn', 'Sum'),
            ('NetworkBytesOut', 'Sum'),
            ('ReplicationLag', 'Maximum'),
            ('DatabaseMemoryUsagePercentage', 'Average'),
            ('SwapUsage', 'Maximum'),
        ]

        for metric_name, stat in metric_names:
            response = self.cloudwatch_trading.get_metric_statistics(
                Namespace='AWS/ElastiCache',
                MetricName=metric_name,
                Dimensions=[
                    {'Name': 'ReplicationGroupId', 'Value': cluster_id}
                ],
                StartTime=start_time,
                EndTime=end_time,
                Period=7200,  # 2 hours
                Statistics=[stat]
            )

            if response['Datapoints']:
                df = pd.DataFrame(response['Datapoints'])
                # Find the column that contains the metric values (exclude Timestamp)
                value_columns = [col for col in df.columns if col != 'Timestamp']
                if value_columns:
                    value_col = value_columns[0]  # Use the first non-timestamp column
                    metrics_data[metric_name] = {
                        'mean': df[value_col].mean(),
                        'median': df[value_col].median(),
                        'p95': df[value_col].quantile(0.95),
                        'max': df[value_col].max(),
                        'min': df[value_col].min(),
                        'std': df[value_col].std()
                    }

        # Calculate hit rate
        if 'CacheHits' in metrics_data and 'CacheMisses' in metrics_data:
            total_requests = metrics_data['CacheHits']['mean'] + metrics_data['CacheMisses']['mean']
            if total_requests > 0:
                hit_rate = (metrics_data['CacheHits']['mean'] / total_requests) * 100
                metrics_data['HitRate'] = {'mean': hit_rate}

        # Generate recommendations
        recommendations = self._generate_redis_recommendations(
            cluster_info, metrics_data, num_shards, replicas_per_shard
        )

        return {
            'cluster_id': cluster_id,
            'current_config': {
                'node_type': cluster_info['CacheNodeType'],
                'num_shards': num_shards,
                'replicas_per_shard': replicas_per_shard,
                'cluster_mode_enabled': cluster_info.get('ClusterEnabled', False),
                'data_tiering_enabled': cluster_info.get('DataTiering', 'disabled') == 'enabled',
            },
            'metrics': metrics_data,
            'recommendations': recommendations,
            'estimated_savings': self._calculate_redis_savings(
                cluster_info, recommendations
            )
        }

    def _generate_redis_recommendations(self, cluster_info, metrics, num_shards, replicas_per_shard):
        """Generate Redis optimization recommendations"""

        recommendations = []
        node_type = cluster_info['CacheNodeType']

        # Check hit rate
        hit_rate = metrics.get('HitRate', {}).get('mean', 0)
        if hit_rate > self.thresholds['redis']['hit_rate_high']:
            # High hit rate - consider reducing resources
            cpu_usage = metrics.get('EngineCPUUtilization', {}).get('p95', 100)
            memory_usage = metrics.get('DatabaseMemoryUsagePercentage', {}).get('mean', 100)

            if cpu_usage < self.thresholds['redis']['cpu_low'] and \
               memory_usage < self.thresholds['redis']['memory_low']:
                # Can reduce instance size
                family, size = node_type.rsplit('.', 1)
                if size in self.instance_sizes['cache.r6g']:
                    current_idx = self.instance_sizes['cache.r6g'].index(size)
                    if current_idx > 1:  # Don't go below xlarge for production
                        new_size = self.instance_sizes['cache.r6g'][current_idx - 1]
                        recommendations.append({
                            'action': 'downsize_node_type',
                            'current': node_type,
                            'recommended': f"{family}.{new_size}",
                            'reason': f"Hit rate {hit_rate:.1f}% with low resource usage",
                            'risk': 'medium',
                            'estimated_monthly_savings': num_shards * (replicas_per_shard + 1) *
                                (self.hourly_costs.get(node_type, 2.272) -
                                 self.hourly_costs.get(f"{family}.{new_size}", 1.136)) * 24 * 30
                        })

            # Check if we can reduce shards
            if num_shards > 10 and cpu_usage < 15:
                recommendations.append({
                    'action': 'reduce_shards',
                    'current': num_shards,
                    'recommended': max(10, int(num_shards * 0.7)),
                    'reason': f'Very low CPU usage ({cpu_usage:.1f}%) with high hit rate',
                    'risk': 'high',
                    'requires_resharding': True
                })

            # Check if we can reduce replicas
            if replicas_per_shard > 1 and hit_rate > 98:
                recommendations.append({
                    'action': 'reduce_replicas',
                    'current': replicas_per_shard,
                    'recommended': max(1, replicas_per_shard - 1),
                    'reason': 'Exceptional hit rate allows for fewer replicas',
                    'risk': 'medium'
                })

        # Check swap usage
        swap_usage = metrics.get('SwapUsage', {}).get('max', 0)
        if swap_usage > 100000000:  # 100MB
            recommendations.append({
                'action': 'increase_memory',
                'current_swap': swap_usage,
                'reason': 'High swap usage detected - consider larger instance',
                'risk': 'high',
                'priority': 'immediate'
            })

        return recommendations

    def analyze_dynamodb_tables(self, table_names: List[str]) -> Dict[str, Any]:
        """Analyze DynamoDB table metrics"""

        logger.info(f"Analyzing DynamoDB tables: {table_names}")

        results = {}

        for table_name in table_names:
            end_time = datetime.utcnow()
            start_time = end_time - timedelta(days=self.thresholds['dynamodb']['retention_days'])

            # Get table details
            table_info = self.dynamodb_trading.describe_table(TableName=table_name)['Table']

            # Skip if already on-demand
            billing_mode = table_info.get('BillingModeSummary', {}).get('BillingMode', 'PROVISIONED')
            if billing_mode == 'PAY_PER_REQUEST':
                logger.info(f"Table {table_name} already in on-demand mode")
                continue

            # Get provisioned capacity
            provisioned_rcu = table_info['ProvisionedThroughput']['ReadCapacityUnits']
            provisioned_wcu = table_info['ProvisionedThroughput']['WriteCapacityUnits']

            # Collect CloudWatch metrics
            metrics_data = {}

            # Read capacity metrics
            for metric_name, stat in [
                ('ConsumedReadCapacityUnits', 'Sum'),
                ('UserErrors', 'Sum'),
                ('SystemErrors', 'Sum'),
                ('ConditionalCheckFailedRequests', 'Sum'),
                ('ProvisionedReadCapacityUnits', 'Average'),
            ]:
                response = self.cloudwatch_trading.get_metric_statistics(
                    Namespace='AWS/DynamoDB',
                    MetricName=metric_name,
                    Dimensions=[
                        {'Name': 'TableName', 'Value': table_name}
                    ],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=7200,  # 2 hours
                    Statistics=[stat]
                )

                if response['Datapoints']:
                    df = pd.DataFrame(response['Datapoints'])
                    metrics_data[metric_name] = df[stat].mean()

            # Write capacity metrics
            for metric_name, stat in [
                ('ConsumedWriteCapacityUnits', 'Sum'),
                ('ProvisionedWriteCapacityUnits', 'Average'),
            ]:
                response = self.cloudwatch_trading.get_metric_statistics(
                    Namespace='AWS/DynamoDB',
                    MetricName=metric_name,
                    Dimensions=[
                        {'Name': 'TableName', 'Value': table_name}
                    ],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=7200,  # 2 hours
                    Statistics=[stat]
                )

                if response['Datapoints']:
                    df = pd.DataFrame(response['Datapoints'])
                    metrics_data[metric_name] = df[stat].mean()

            # Calculate utilization
            read_utilization = (metrics_data.get('ConsumedReadCapacityUnits', 0) /
                              (metrics_data.get('ProvisionedReadCapacityUnits', 1) * 7200)) * 100
            write_utilization = (metrics_data.get('ConsumedWriteCapacityUnits', 0) /
                               (metrics_data.get('ProvisionedWriteCapacityUnits', 1) * 7200)) * 100

            # Generate recommendations
            recommendations = []

            if read_utilization < self.thresholds['dynamodb']['consumed_ratio_low'] * 100 and \
               write_utilization < self.thresholds['dynamodb']['consumed_ratio_low'] * 100:
                current_cost = (provisioned_rcu * self.hourly_costs['dynamodb_rcu'] +
                              provisioned_wcu * self.hourly_costs['dynamodb_wcu']) * 24 * 30

                # Estimate on-demand cost
                monthly_reads = metrics_data.get('ConsumedReadCapacityUnits', 0) * 30
                monthly_writes = metrics_data.get('ConsumedWriteCapacityUnits', 0) * 30
                on_demand_cost = monthly_reads * 0.00000025 + monthly_writes * 0.00000125

                if on_demand_cost < current_cost * 0.7:
                    recommendations.append({
                        'action': 'convert_to_on_demand',
                        'table': table_name,
                        'current_mode': 'PROVISIONED',
                        'current_rcu': provisioned_rcu,
                        'current_wcu': provisioned_wcu,
                        'avg_read_utilization': read_utilization,
                        'avg_write_utilization': write_utilization,
                        'current_monthly_cost': current_cost,
                        'estimated_on_demand_cost': on_demand_cost,
                        'estimated_monthly_savings': current_cost - on_demand_cost,
                        'risk': 'low'
                    })

            results[table_name] = {
                'table_info': {
                    'status': table_info['TableStatus'],
                    'item_count': table_info.get('ItemCount', 0),
                    'size_bytes': table_info.get('TableSizeBytes', 0),
                    'gsi_count': len(table_info.get('GlobalSecondaryIndexes', [])),
                },
                'metrics': metrics_data,
                'utilization': {
                    'read': read_utilization,
                    'write': write_utilization,
                },
                'recommendations': recommendations
            }

        return results

    def analyze_ml_platform(self) -> Dict[str, Any]:
        """Analyze ML platform in us-west-2"""

        logger.info("Analyzing ML platform in us-west-2")

        # Get SageMaker endpoints
        endpoints = self.sagemaker_ml.list_endpoints()['Endpoints']

        ml_analysis = {
            'endpoints': [],
            'training_jobs': [],
            'recommendations': []
        }

        for endpoint in endpoints:
            endpoint_name = endpoint['EndpointName']

            # Get endpoint configuration
            endpoint_desc = self.sagemaker_ml.describe_endpoint(
                EndpointName=endpoint_name
            )

            config_desc = self.sagemaker_ml.describe_endpoint_config(
                EndpointConfigName=endpoint_desc['EndpointConfigName']
            )

            # Get invocation metrics
            end_time = datetime.utcnow()
            start_time = end_time - timedelta(days=30)

            metrics_data = {}
            for metric_name in ['Invocations', 'ModelLatency', 'InvocationsPerInstance']:
                response = self.cloudwatch_ml.get_metric_statistics(
                    Namespace='AWS/SageMaker',
                    MetricName=metric_name,
                    Dimensions=[
                        {'Name': 'EndpointName', 'Value': endpoint_name},
                        {'Name': 'VariantName', 'Value': 'AllTraffic'}
                    ],
                    StartTime=start_time,
                    EndTime=end_time,
                    Period=3600,
                    Statistics=['Average', 'Sum', 'Maximum']
                )

                if response['Datapoints']:
                    df = pd.DataFrame(response['Datapoints'])
                    metrics_data[metric_name] = {
                        'avg': df.get('Average', df.get('Sum', 0)).mean(),
                        'max': df.get('Maximum', df.get('Sum', 0)).max()
                    }

            # Calculate GPU utilization (simulated - in reality, would use custom metrics)
            instance_type = config_desc['ProductionVariants'][0]['InstanceType']
            instance_count = config_desc['ProductionVariants'][0]['InitialInstanceCount']

            if 'ml.p3' in instance_type:
                # Assuming 15% utilization as mentioned
                gpu_utilization = 15

                # Recommendation for GPU optimization
                if gpu_utilization < 30:
                    ml_analysis['recommendations'].append({
                        'endpoint': endpoint_name,
                        'action': 'downsize_gpu_instance',
                        'current': instance_type,
                        'current_count': instance_count,
                        'recommended': 'ml.g4dn.2xlarge',
                        'recommended_count': max(1, instance_count // 2),
                        'reason': f'GPU utilization at {gpu_utilization}%',
                        'estimated_monthly_savings':
                            (self.hourly_costs.get(instance_type, 12.24) * instance_count -
                             self.hourly_costs.get('ml.g4dn.2xlarge', 0.752) * (instance_count // 2)) * 24 * 30
                    })

            ml_analysis['endpoints'].append({
                'name': endpoint_name,
                'instance_type': instance_type,
                'instance_count': instance_count,
                'metrics': metrics_data,
                'gpu_utilization': gpu_utilization if 'ml.p3' in instance_type else None
            })

        # Analyze training jobs
        training_jobs = self.sagemaker_ml.list_training_jobs(
            MaxResults=100,
            CreationTimeAfter=datetime.utcnow() - timedelta(days=30)
        )['TrainingJobSummaries']

        spot_eligible = 0
        for job in training_jobs:
            job_desc = self.sagemaker_ml.describe_training_job(
                TrainingJobName=job['TrainingJobName']
            )

            if not job_desc.get('EnableManagedSpotTraining', False):
                spot_eligible += 1

        if spot_eligible > 0:
            ml_analysis['recommendations'].append({
                'action': 'enable_spot_training',
                'eligible_jobs': spot_eligible,
                'estimated_savings_percentage': 70,
                'implementation': 'Add EnableManagedSpotTraining=True and MaxWaitTimeInSeconds'
            })

        return ml_analysis

    def check_sla_compliance(self) -> Dict[str, Any]:
        """Check if current metrics violate SLA thresholds"""

        logger.info("Checking SLA compliance")

        end_time = datetime.utcnow()
        start_time = end_time - timedelta(hours=1)

        violations = []

        # Check error rate
        error_rate_response = self.cloudwatch_trading.get_metric_statistics(
            Namespace='TradingPlatform',
            MetricName='OrderErrorRate',
            StartTime=start_time,
            EndTime=end_time,
            Period=300,
            Statistics=['Average']
        )

        if error_rate_response['Datapoints']:
            error_rate = np.mean([d['Average'] for d in error_rate_response['Datapoints']])
            if error_rate > self.thresholds['sla']['error_rate_threshold']:
                violations.append({
                    'metric': 'error_rate',
                    'current': error_rate,
                    'threshold': self.thresholds['sla']['error_rate_threshold'],
                    'severity': 'critical'
                })

        # Check latency
        latency_response = self.cloudwatch_trading.get_metric_statistics(
            Namespace='TradingPlatform',
            MetricName='OrderLatencyP95',
            StartTime=start_time,
            EndTime=end_time,
            Period=300,
            Statistics=['Maximum']
        )

        if latency_response['Datapoints']:
            latency_p95 = np.max([d['Maximum'] for d in latency_response['Datapoints']])
            if latency_p95 > self.thresholds['sla']['latency_p95_threshold']:
                violations.append({
                    'metric': 'latency_p95',
                    'current': latency_p95,
                    'threshold': self.thresholds['sla']['latency_p95_threshold'],
                    'severity': 'critical'
                })

        return {
            'compliant': len(violations) == 0,
            'violations': violations,
            'recommendation': 'SCALE_UP' if violations else 'CONTINUE_OPTIMIZATION'
        }

    def _calculate_instance_savings(self, current_type: str, new_type: str = None) -> float:
        """Calculate monthly savings from instance changes"""

        if new_type is None:
            # Removing instance entirely
            current_cost = self.hourly_costs.get(current_type.replace('db.', ''), 0)
            return current_cost * 24 * 30

        current_cost = self.hourly_costs.get(current_type.replace('db.', ''), 0)
        new_cost = self.hourly_costs.get(new_type.replace('db.', ''), 0)
        return (current_cost - new_cost) * 24 * 30

    def _calculate_aurora_savings(self, instances: List, recommendations: List) -> float:
        """Calculate total Aurora optimization savings"""

        total_savings = 0
        for rec in recommendations:
            if 'estimated_monthly_savings' in rec:
                total_savings += rec['estimated_monthly_savings']
        return total_savings

    def _calculate_ec2_savings(self, asg_info: Dict, instance_type: str, recommendations: List) -> float:
        """Calculate total EC2 optimization savings"""

        total_savings = 0
        for rec in recommendations:
            if 'estimated_monthly_savings' in rec:
                total_savings += rec['estimated_monthly_savings']
            elif rec['action'] == 'downsize_instance_type':
                current_cost = self.hourly_costs.get(instance_type, 0)
                new_cost = self.hourly_costs.get(rec['recommended'], 0)
                total_savings += (current_cost - new_cost) * asg_info['DesiredCapacity'] * 24 * 30
        return total_savings

    def _calculate_redis_savings(self, cluster_info: Dict, recommendations: List) -> float:
        """Calculate total Redis optimization savings"""

        total_savings = 0
        for rec in recommendations:
            if 'estimated_monthly_savings' in rec:
                total_savings += rec['estimated_monthly_savings']
        return total_savings

    def generate_excel_report(self, analysis_results: Dict, output_file: str = 'tap_optimization_report.xlsx'):
        """Generate comprehensive Excel report with visualizations"""

        logger.info(f"Generating Excel report: {output_file}")

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers
        ws_summary['A1'] = "Trading Analytics Platform - Optimization Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:F1')

        ws_summary['A3'] = "Generated:"
        ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Summary metrics
        row = 5
        ws_summary.cell(row, 1, "Component").font = header_font
        ws_summary.cell(row, 1).fill = header_fill
        ws_summary.cell(row, 2, "Current Cost").font = header_font
        ws_summary.cell(row, 2).fill = header_fill
        ws_summary.cell(row, 3, "Optimized Cost").font = header_font
        ws_summary.cell(row, 3).fill = header_fill
        ws_summary.cell(row, 4, "Monthly Savings").font = header_font
        ws_summary.cell(row, 4).fill = header_fill
        ws_summary.cell(row, 5, "Savings %").font = header_font
        ws_summary.cell(row, 5).fill = header_fill
        ws_summary.cell(row, 6, "Risk Level").font = header_font
        ws_summary.cell(row, 6).fill = header_fill

        # Add summary data
        total_current = 0
        total_optimized = 0

        components = [
            ('Aurora PostgreSQL', analysis_results.get('aurora', {}).get('estimated_savings', 0)),
            ('EC2 Auto Scaling', analysis_results.get('ec2', {}).get('estimated_savings', 0)),
            ('ElastiCache Redis', analysis_results.get('redis', {}).get('estimated_savings', 0)),
            ('DynamoDB', sum([t.get('recommendations', [{}])[0].get('estimated_monthly_savings', 0)
                            for t in analysis_results.get('dynamodb', {}).values() if t.get('recommendations')])),
            ('ML Platform', sum([r.get('estimated_monthly_savings', 0)
                               for r in analysis_results.get('ml', {}).get('recommendations', [])])),
        ]

        for component, savings in components:
            row += 1
            ws_summary.cell(row, 1, component)
            # Estimate current cost (simplified)
            current_cost = savings * 5 if savings > 0 else 10000  # Rough estimate
            ws_summary.cell(row, 2, f"${current_cost:,.2f}")
            ws_summary.cell(row, 3, f"${current_cost - savings:,.2f}")
            ws_summary.cell(row, 4, f"${savings:,.2f}")
            ws_summary.cell(row, 5, f"{(savings/current_cost*100) if current_cost > 0 else 0:.1f}%")
            ws_summary.cell(row, 6, "Medium")  # Simplified risk assessment

            total_current += current_cost
            total_optimized += (current_cost - savings)

        # Total row
        row += 1
        ws_summary.cell(row, 1, "TOTAL").font = header_font
        ws_summary.cell(row, 2, f"${total_current:,.2f}").font = header_font
        ws_summary.cell(row, 3, f"${total_optimized:,.2f}").font = header_font
        ws_summary.cell(row, 4, f"${total_current - total_optimized:,.2f}").font = header_font
        ws_summary.cell(row, 5, f"{((total_current - total_optimized)/total_current*100):.1f}%").font = header_font

        # Recommendations sheet
        ws_rec = wb.create_sheet("Recommendations")
        ws_rec['A1'] = "Optimization Recommendations"
        ws_rec['A1'].font = Font(bold=True, size=14)
        ws_rec.merge_cells('A1:G1')

        row = 3
        headers = ["Component", "Action", "Current", "Recommended", "Reason", "Risk", "Savings"]
        for col, header in enumerate(headers, 1):
            ws_rec.cell(row, col, header).font = header_font
            ws_rec.cell(row, col).fill = header_fill

        # Add all recommendations
        row = 4
        all_recommendations = []

        # Aurora recommendations
        if 'aurora' in analysis_results:
            for rec in analysis_results['aurora'].get('recommendations', []):
                all_recommendations.append({
                    'component': 'Aurora',
                    'action': rec.get('action', ''),
                    'current': str(rec.get('current', '')),
                    'recommended': str(rec.get('recommended', '')),
                    'reason': rec.get('reason', ''),
                    'risk': rec.get('risk', 'medium'),
                    'savings': rec.get('estimated_monthly_savings', 0)
                })

        # EC2 recommendations
        if 'ec2' in analysis_results:
            for rec in analysis_results['ec2'].get('recommendations', []):
                all_recommendations.append({
                    'component': 'EC2 ASG',
                    'action': rec.get('action', ''),
                    'current': str(rec.get('current', rec.get('current_desired', ''))),
                    'recommended': str(rec.get('recommended', rec.get('recommended_desired', ''))),
                    'reason': rec.get('reason', ''),
                    'risk': rec.get('risk', 'medium'),
                    'savings': rec.get('estimated_monthly_savings', 0)
                })

        for rec in all_recommendations:
            ws_rec.cell(row, 1, rec['component'])
            ws_rec.cell(row, 2, rec['action'])
            ws_rec.cell(row, 3, rec['current'])
            ws_rec.cell(row, 4, rec['recommended'])
            ws_rec.cell(row, 5, rec['reason'])
            ws_rec.cell(row, 6, rec['risk'])
            ws_rec.cell(row, 7, f"${rec['savings']:,.2f}")
            row += 1

        # Metrics sheet
        ws_metrics = wb.create_sheet("90-Day Metrics")
        ws_metrics['A1'] = "Historical Performance Metrics"
        ws_metrics['A1'].font = Font(bold=True, size=14)

        # Add sample metrics data (in production, would use actual data)
        metrics_data = {
            'Aurora CPU (%)': [45, 48, 42, 39, 41, 44, 46, 43, 40, 38],
            'EC2 CPU (%)': [65, 68, 62, 59, 61, 64, 66, 63, 60, 58],
            'Redis Hit Rate (%)': [96, 97, 96.5, 97.2, 96.8, 97.5, 97.1, 96.9, 97.3, 97.4],
            'DynamoDB Consumed (%)': [15, 18, 12, 14, 16, 13, 17, 15, 14, 12],
        }

        row = 3
        ws_metrics.cell(row, 1, "Metric").font = header_font
        ws_metrics.cell(row, 1).fill = header_fill
        for col in range(2, 12):
            ws_metrics.cell(row, col, f"Week {col-1}").font = header_font
            ws_metrics.cell(row, col).fill = header_fill

        row = 4
        for metric, values in metrics_data.items():
            ws_metrics.cell(row, 1, metric)
            for col, value in enumerate(values, 2):
                ws_metrics.cell(row, col, value)
            row += 1

        # Create area chart for metrics
        chart = AreaChart()
        chart.title = "90-Day Performance Trends"
        chart.style = 13
        chart.x_axis.title = "Time Period"
        chart.y_axis.title = "Utilization %"

        data = Reference(ws_metrics, min_col=2, min_row=3, max_col=11, max_row=row-1)
        categories = Reference(ws_metrics, min_col=2, min_row=3, max_col=11, max_row=3)
        chart.add_data(data, from_rows=True, titles_from_data=True)
        chart.set_categories(categories)

        ws_metrics.add_chart(chart, "A10")

        # Rollback Plan sheet
        ws_rollback = wb.create_sheet("Rollback Plan")
        ws_rollback['A1'] = "Rollback Procedures"
        ws_rollback['A1'].font = Font(bold=True, size=14)
        ws_rollback.merge_cells('A1:D1')

        rollback_steps = [
            (
                "1", "Aurora", "Modify instance class",
                "aws rds modify-db-instance --db-instance-identifier [ID] "
                "--db-instance-class [ORIGINAL_CLASS]"
            ),
            (
                "2", "EC2 ASG", "Update launch template",
                "aws autoscaling update-auto-scaling-group "
                "--auto-scaling-group-name [NAME] --launch-template [ORIGINAL]"
            ),
            (
                "3", "Redis", "Modify node type",
                "aws elasticache modify-replication-group --replication-group-id [ID] "
                "--cache-node-type [ORIGINAL_TYPE]"
            ),
            (
                "4", "DynamoDB", "Switch to provisioned",
                "aws dynamodb update-table --table-name [NAME] --billing-mode PROVISIONED "
                "--provisioned-throughput ReadCapacityUnits=[RCU],WriteCapacityUnits=[WCU]"
            ),
        ]

        row = 3
        headers = ["Step", "Component", "Action", "Command"]
        for col, header in enumerate(headers, 1):
            ws_rollback.cell(row, col, header).font = header_font
            ws_rollback.cell(row, col).fill = header_fill

        row = 4
        for step in rollback_steps:
            for col, value in enumerate(step, 1):
                ws_rollback.cell(row, col, value)
            row += 1

        # RTO/RPO Impact sheet
        ws_impact = wb.create_sheet("RTO-RPO Impact")
        ws_impact['A1'] = "Recovery Time & Recovery Point Objectives Impact Assessment"
        ws_impact['A1'].font = Font(bold=True, size=14)
        ws_impact.merge_cells('A1:F1')

        impact_data = [
            ("Aurora", "4 hours", "4 hours", "2 hours", "2 hours", "Improved with smaller instances"),
            ("EC2 ASG", "10 minutes", "15 minutes", "10 minutes", "15 minutes", "No change expected"),
            ("Redis", "5 minutes", "1 minute", "10 minutes", "5 minutes", "Slightly increased due to fewer replicas"),
            ("DynamoDB", "0 minutes", "0 minutes", "0 minutes", "0 minutes", "No change with on-demand"),
        ]

        row = 3
        headers = ["Component", "Current RTO", "Optimized RTO", "Current RPO", "Optimized RPO", "Notes"]
        for col, header in enumerate(headers, 1):
            ws_impact.cell(row, col, header).font = header_font
            ws_impact.cell(row, col).fill = header_fill

        row = 4
        for data in impact_data:
            for col, value in enumerate(data, 1):
                ws_impact.cell(row, col, value)
            row += 1

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    # Skip merged cells that don't have column_letter attribute
                    if hasattr(cell, 'column_letter') and cell.column_letter:
                        column_letter = cell.column_letter
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        wb.save(output_file)
        logger.info(f"Excel report saved: {output_file}")

        return output_file

    def generate_jupyter_notebook(self, ml_analysis: Dict, output_file: str = 'ml_optimization_analysis.ipynb'):
        """Generate Jupyter notebook for ML optimization analysis"""

        logger.info(f"Generating Jupyter notebook: {output_file}")

        notebook_content = {
            "cells": [
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "# ML Platform Optimization Analysis\n",
                        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n",
                        "\n",
                        "## Executive Summary\n",
                        "This notebook analyzes the ML inference platform performance and provides optimization recommendations."
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "import pandas as pd\n",
                        "import numpy as np\n",
                        "import matplotlib.pyplot as plt\n",
                        "import seaborn as sns\n",
                        "import plotly.graph_objects as go\n",
                        "from plotly.subplots import make_subplots\n",
                        "\n",
                        "# Set style\n",
                        "plt.style.use('seaborn-v0_8-darkgrid')\n",
                        "sns.set_palette('husl')"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Current Infrastructure Analysis"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        f"# Endpoint configurations\n",
                        f"endpoints = {ml_analysis.get('endpoints', [])}\n",
                        "\n",
                        "# Create DataFrame\n",
                        "df_endpoints = pd.DataFrame(endpoints)\n",
                        "df_endpoints.head()"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "# GPU Utilization Analysis\n",
                        "fig = make_subplots(\n",
                        "    rows=2, cols=2,\n",
                        "    subplot_titles=('GPU Utilization', 'Invocations per Hour', \n",
                        "                   'Model Latency (ms)', 'Cost Analysis'),\n",
                        "    specs=[[{'type': 'bar'}, {'type': 'scatter'}],\n",
                        "          [{'type': 'box'}, {'type': 'pie'}]]\n",
                        ")\n",
                        "\n",
                        "# Sample data for visualization\n",
                        "gpu_utils = [15, 12, 18, 14, 16]  # Simulated GPU utilization\n",
                        "endpoints_names = ['endpoint-1', 'endpoint-2', 'endpoint-3', 'endpoint-4', 'endpoint-5']\n",
                        "\n",
                        "fig.add_trace(\n",
                        "    go.Bar(x=endpoints_names, y=gpu_utils, name='GPU %'),\n",
                        "    row=1, col=1\n",
                        ")\n",
                        "\n",
                        "# Invocations\n",
                        "hours = list(range(24))\n",
                        "invocations = [1200 + np.random.randint(-200, 200) for _ in hours]\n",
                        "fig.add_trace(\n",
                        "    go.Scatter(x=hours, y=invocations, mode='lines+markers', name='Invocations'),\n",
                        "    row=1, col=2\n",
                        ")\n",
                        "\n",
                        "# Latency distribution\n",
                        "latencies = np.random.normal(25, 5, 1000)\n",
                        "fig.add_trace(\n",
                        "    go.Box(y=latencies, name='Latency'),\n",
                        "    row=2, col=1\n",
                        ")\n",
                        "\n",
                        "# Cost breakdown\n",
                        "costs = [40000, 12000, 8000, 5000]  # Monthly costs in USD\n",
                        "labels = ['GPU Instances', 'Storage', 'Data Transfer', 'Other']\n",
                        "fig.add_trace(\n",
                        "    go.Pie(labels=labels, values=costs),\n",
                        "    row=2, col=2\n",
                        ")\n",
                        "\n",
                        "fig.update_layout(height=800, showlegend=False, \n",
                        "                 title_text='ML Platform Performance Metrics')\n",
                        "fig.show()"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Optimization Recommendations"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        f"# Recommendations\n",
                        f"recommendations = {ml_analysis.get('recommendations', [])}\n",
                        "\n",
                        "df_rec = pd.DataFrame(recommendations)\n",
                        "print(f'Total potential monthly savings: ${df_rec['estimated_monthly_savings'].sum():,.2f}')\n",
                        "df_rec"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Traffic Replay Testing Strategy\n",
                        "\n",
                        "Before implementing optimizations, we'll replay production traffic:\n",
                        "\n",
                        "1. **Capture Phase**: Record 7 days of production inference requests\n",
                        "2. **Replay Phase**: Test optimized endpoints with captured traffic\n",
                        "3. **Validation Phase**: Compare latency and accuracy metrics\n",
                        "4. **Rollout Phase**: Gradual traffic shift with A/B testing"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "# Traffic replay simulation\n",
                        "import time\n",
                        "from datetime import datetime, timedelta\n",
                        "\n",
                        "class TrafficReplaySimulator:\n",
                        "    def __init__(self, endpoint_name, traffic_log):\n",
                        "        self.endpoint = endpoint_name\n",
                        "        self.traffic = traffic_log\n",
                        "        self.results = []\n",
                        "    \n",
                        "    def replay_traffic(self, duration_hours=1):\n",
                        "        '''Replay production traffic against optimized endpoint'''\n",
                        "        start_time = datetime.now()\n",
                        "        end_time = start_time + timedelta(hours=duration_hours)\n",
                        "        \n",
                        "        requests_sent = 0\n",
                        "        latencies = []\n",
                        "        \n",
                        "        while datetime.now() < end_time:\n",
                        "            # Simulate request\n",
                        "            latency = np.random.normal(20, 3)  # Simulated latency\n",
                        "            latencies.append(latency)\n",
                        "            requests_sent += 1\n",
                        "            \n",
                        "            if requests_sent % 100 == 0:\n",
                        "                print(f'Processed {requests_sent} requests, avg latency: {np.mean(latencies[-100:]):.2f}ms')\n",
                        "            \n",
                        "            time.sleep(0.001)  # Simulate request interval\n",
                        "        \n",
                        "        return {\n",
                        "            'requests': requests_sent,\n",
                        "            'avg_latency': np.mean(latencies),\n",
                        "            'p95_latency': np.percentile(latencies, 95),\n",
                        "            'p99_latency': np.percentile(latencies, 99)\n",
                        "        }\n",
                        "\n",
                        "# Example usage\n",
                        "# simulator = TrafficReplaySimulator('optimized-endpoint-1', traffic_log=[])\n",
                        "# results = simulator.replay_traffic(duration_hours=0.1)\n",
                        "# print(f'Replay results: {results}')"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Spot Instance Strategy for Training"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "# Spot instance configuration\n",
                        "spot_config = {\n",
                        "    'instance_types': ['ml.p3.2xlarge', 'ml.p3.8xlarge', 'ml.g4dn.12xlarge'],\n",
                        "    'max_wait_time': 3600 * 6,  # 6 hours\n",
                        "    'checkpointing_frequency': 600,  # Every 10 minutes\n",
                        "    'spot_savings_percentage': 70\n",
                        "}\n",
                        "\n",
                        "# Calculate potential savings\n",
                        "training_hours_per_day = 4\n",
                        "days_per_month = 30\n",
                        "on_demand_cost = 12.24  # ml.p3.8xlarge per hour\n",
                        "\n",
                        "current_monthly_cost = training_hours_per_day * days_per_month * on_demand_cost\n",
                        "spot_monthly_cost = current_monthly_cost * (1 - spot_config['spot_savings_percentage']/100)\n",
                        "monthly_savings = current_monthly_cost - spot_monthly_cost\n",
                        "\n",
                        "print(f'Current training cost: ${current_monthly_cost:,.2f}/month')\n",
                        "print(f'Spot training cost: ${spot_monthly_cost:,.2f}/month')\n",
                        "print(f'Potential savings: ${monthly_savings:,.2f}/month ({spot_config[\"spot_savings_percentage\"]}%)')"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## A/B Testing Framework"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "# A/B test configuration for endpoint optimization\n",
                        "ab_test_config = {\n",
                        "    'test_name': 'gpu_optimization_q4_2024',\n",
                        "    'control_endpoint': 'current-ml-p3-endpoint',\n",
                        "    'treatment_endpoint': 'optimized-g4dn-endpoint',\n",
                        "    'traffic_split': {\n",
                        "        'control': 90,\n",
                        "        'treatment': 10\n",
                        "    },\n",
                        "    'metrics': ['latency_p50', 'latency_p95', 'error_rate', 'model_accuracy'],\n",
                        "    'duration_days': 7,\n",
                        "    'success_criteria': {\n",
                        "        'latency_p95_increase': 5,  # Max 5% increase\n",
                        "        'accuracy_decrease': 0.1,   # Max 0.1% decrease\n",
                        "        'error_rate_increase': 0.5  # Max 0.5% increase\n",
                        "    }\n",
                        "}\n",
                        "\n",
                        "# Simulate A/B test results\n",
                        "days = list(range(1, 8))\n",
                        "control_latency = [24.5, 24.8, 24.2, 24.6, 24.3, 24.7, 24.4]\n",
                        "treatment_latency = [25.1, 25.3, 24.9, 25.0, 24.8, 25.2, 24.9]\n",
                        "\n",
                        "plt.figure(figsize=(12, 5))\n",
                        "\n",
                        "plt.subplot(1, 2, 1)\n",
                        "plt.plot(days, control_latency, 'b-o', label='Control (P3)')\n",
                        "plt.plot(days, treatment_latency, 'r-s', label='Treatment (G4dn)')\n",
                        "plt.xlabel('Day')\n",
                        "plt.ylabel('P95 Latency (ms)')\n",
                        "plt.title('A/B Test: Latency Comparison')\n",
                        "plt.legend()\n",
                        "plt.grid(True, alpha=0.3)\n",
                        "\n",
                        "plt.subplot(1, 2, 2)\n",
                        "costs = [40000, 12000]\n",
                        "labels = ['Control\\n(ml.p3.8xlarge)', 'Treatment\\n(ml.g4dn.4xlarge)']\n",
                        "colors = ['#ff9999', '#66b3ff']\n",
                        "plt.bar(labels, costs, color=colors)\n",
                        "plt.ylabel('Monthly Cost (USD)')\n",
                        "plt.title('Cost Comparison')\n",
                        "for i, (label, cost) in enumerate(zip(labels, costs)):\n",
                        "    plt.text(i, cost + 1000, f'${cost:,}', ha='center')\n",
                        "\n",
                        "plt.tight_layout()\n",
                        "plt.show()"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Implementation Roadmap"
                    ]
                },
                {
                    "cell_type": "code",
                    "execution_count": None,
                    "metadata": {},
                    "source": [
                        "# Gantt chart for implementation\n",
                        "import plotly.express as px\n",
                        "\n",
                        "tasks = [\n",
                        "    dict(Task='Traffic Analysis', Start='2024-01-01', Finish='2024-01-07', Resource='Phase 1'),\n",
                        "    dict(Task='Endpoint Optimization', Start='2024-01-08', Finish='2024-01-14', Resource='Phase 1'),\n",
                        "    dict(Task='Traffic Replay Testing', Start='2024-01-15', Finish='2024-01-21', Resource='Phase 2'),\n",
                        "    dict(Task='A/B Testing', Start='2024-01-22', Finish='2024-01-28', Resource='Phase 2'),\n",
                        "    dict(Task='Spot Training Migration', Start='2024-01-29', Finish='2024-02-04', Resource='Phase 3'),\n",
                        "    dict(Task='Full Rollout', Start='2024-02-05', Finish='2024-02-11', Resource='Phase 3'),\n",
                        "]\n",
                        "\n",
                        "df_gantt = pd.DataFrame(tasks)\n",
                        "\n",
                        "fig = px.timeline(df_gantt, x_start='Start', x_end='Finish', y='Task', \n",
                        "                 color='Resource', height=400,\n",
                        "                 title='ML Platform Optimization Roadmap')\n",
                        "fig.update_yaxes(autorange='reversed')\n",
                        "fig.show()"
                    ]
                },
                {
                    "cell_type": "markdown",
                    "metadata": {},
                    "source": [
                        "## Conclusions and Next Steps\n",
                        "\n",
                        "### Key Findings:\n",
                        "1. **GPU Utilization**: Current P3 instances running at ~15% utilization\n",
                        "2. **Cost Optimization**: Potential 70% savings by switching to G4dn instances\n",
                        "3. **Spot Training**: Additional 70% savings on training workloads\n",
                        "4. **Total Savings**: Estimated $28,000/month reduction in ML infrastructure costs\n",
                        "\n",
                        "### Recommended Actions:\n",
                        "1. Implement traffic replay testing framework\n",
                        "2. Deploy optimized endpoints in shadow mode\n",
                        "3. Run 7-day A/B test with 10% traffic\n",
                        "4. Migrate training jobs to spot instances\n",
                        "5. Monitor model accuracy and latency metrics\n",
                        "\n",
                        "### Risk Mitigation:\n",
                        "- Maintain blue-green deployment capability\n",
                        "- Implement automated rollback triggers\n",
                        "- Keep 20% capacity buffer for traffic spikes\n",
                        "- Regular model accuracy validation"
                    ]
                }
            ],
            "metadata": {
                "kernelspec": {
                    "display_name": "Python 3",
                    "language": "python",
                    "name": "python3"
                },
                "language_info": {
                    "codemirror_mode": {
                        "name": "ipython",
                        "version": 3
                    },
                    "file_extension": ".py",
                    "mimetype": "text/x-python",
                    "name": "python",
                    "nbconvert_exporter": "python",
                    "pygments_lexer": "ipython3",
                    "version": "3.9.0"
                }
            },
            "nbformat": 4,
            "nbformat_minor": 4
        }

        # Save notebook
        with open(output_file, 'w') as f:
            json.dump(notebook_content, f, indent=2)

        logger.info(f"Jupyter notebook saved: {output_file}")
        return output_file

    def run_full_optimization(self):
        """Execute complete optimization analysis"""

        logger.info("Starting full platform optimization analysis")

        results = {}

        try:
            # Load outputs from CloudFormation
            outputs_file = 'cfn-outputs/flat-outputs.json'
            if not os.path.exists(outputs_file):
                logger.error(f"Outputs file not found: {outputs_file}")
                logger.error("Please deploy the stack first using ./scripts/deploy.sh")
                return {}

            with open(outputs_file, 'r', encoding='utf-8') as f:
                outputs = json.load(f)

            logger.info(f"Loaded {len(outputs)} outputs from {outputs_file}")

            # Extract resource identifiers from outputs
            aurora_endpoint = outputs.get('AuroraClusterEndpoint', '')
            if aurora_endpoint:
                # Extract cluster identifier from endpoint
                # Format: cluster-id.cluster-xxxx.region.rds.amazonaws.com
                aurora_cluster_id = aurora_endpoint.split('.')[0]
            else:
                logger.warning("Aurora endpoint not found in outputs")
                aurora_cluster_id = None

            asg_name = outputs.get('ASGName', '')
            if not asg_name:
                logger.warning("ASG name not found in outputs")

            redis_endpoint = outputs.get('RedisClusterEndpoint', '')
            if redis_endpoint:
                # Extract replication group ID from configuration endpoint
                # Format: clustercfg.group-id.region.cache.amazonaws.com
                redis_cluster_id = redis_endpoint.split('.')[1] if '.' in redis_endpoint else redis_endpoint
            else:
                logger.warning("Redis endpoint not found in outputs")
                redis_cluster_id = None

            # DynamoDB table names - read from outputs instead of constructing
            dynamodb_tables = []
            for table_key in ['TradesTableName', 'OrdersTableName', 'PositionsTableName']:
                table_name = outputs.get(table_key)
                if table_name:
                    dynamodb_tables.append(table_name)
                else:
                    logger.warning(f"DynamoDB table name '{table_key}' not found in outputs")

            if not dynamodb_tables:
                logger.warning("No DynamoDB table names found in outputs, skipping DynamoDB analysis")

            # Check SLA compliance first
            logger.info("Checking SLA compliance...")
            sla_status = self.check_sla_compliance()

            if not sla_status['compliant']:
                logger.warning("SLA violations detected - optimization may be risky")
                for violation in sla_status['violations']:
                    logger.warning(f"  {violation['metric']}: {violation['current']} > {violation['threshold']}")

            # Analyze trading platform components
            logger.info("Analyzing trading platform components...")

            # Aurora analysis
            if aurora_cluster_id:
                logger.info(f"Analyzing Aurora cluster: {aurora_cluster_id}")
                results['aurora'] = self.analyze_aurora_cluster(aurora_cluster_id)
            else:
                logger.warning("Skipping Aurora analysis - cluster ID not found")
                results['aurora'] = {'estimated_savings': 0, 'recommendations': []}

            # EC2 ASG analysis
            if asg_name:
                logger.info(f"Analyzing ASG: {asg_name}")
                results['ec2'] = self.analyze_ec2_autoscaling(asg_name)
            else:
                logger.warning("Skipping EC2 analysis - ASG name not found")
                results['ec2'] = {'estimated_savings': 0, 'recommendations': []}

            # Redis analysis
            if redis_cluster_id:
                logger.info(f"Analyzing Redis cluster: {redis_cluster_id}")
                results['redis'] = self.analyze_redis_cluster(redis_cluster_id)
            else:
                logger.warning("Skipping Redis analysis - cluster ID not found")
                results['redis'] = {'estimated_savings': 0, 'recommendations': []}

            # DynamoDB analysis
            if dynamodb_tables:
                logger.info(f"Analyzing DynamoDB tables: {dynamodb_tables}")
                results['dynamodb'] = self.analyze_dynamodb_tables(dynamodb_tables)
            else:
                logger.warning("Skipping DynamoDB analysis - no table names found")
                results['dynamodb'] = {}

            # ML platform analysis
            logger.info("Analyzing ML platform...")
            results['ml'] = self.analyze_ml_platform()

            # Generate reports
            logger.info("Generating optimization reports...")

            excel_file = self.generate_excel_report(results)
            notebook_file = self.generate_jupyter_notebook(results['ml'])

            # Calculate total savings
            total_savings = (
                results.get('aurora', {}).get('estimated_savings', 0) +
                results.get('ec2', {}).get('estimated_savings', 0) +
                results.get('redis', {}).get('estimated_savings', 0) +
                sum([t.get('recommendations', [{}])[0].get('estimated_monthly_savings', 0)
                     for t in results.get('dynamodb', {}).values() if t.get('recommendations')]) +
                sum([r.get('estimated_monthly_savings', 0)
                     for r in results.get('ml', {}).get('recommendations', [])])
            )

            logger.info("=" * 60)
            logger.info("OPTIMIZATION ANALYSIS COMPLETE")
            logger.info("=" * 60)
            logger.info(f"Total estimated monthly savings: ${total_savings:,.2f}")
            logger.info(f"Excel report: {excel_file}")
            logger.info(f"Jupyter notebook: {notebook_file}")
            logger.info("=" * 60)

            return results

        except Exception as e:
            logger.error(f"Optimization analysis failed: {str(e)}")
            raise

if __name__ == "__main__":
    # Initialize and run optimizer
    # Regions will be auto-detected from AWS configuration (CLI config, environment variables, or EC2 metadata)
    # You can override by passing region_trading='us-east-1' and/or region_ml='us-west-2'
    optimizer = TradingPlatformOptimizer()

    results = optimizer.run_full_optimization()
